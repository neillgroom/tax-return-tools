#!/usr/bin/env python3
"""
Checksheet Populator
====================
Takes parsed CCH data (JSON) and populates the tax verification checksheet.

Usage:
    python populate_checksheet.py parsed_data.json
    python populate_checksheet.py parsed_data.json --checksheet blank_checksheet.xlsx
    python populate_checksheet.py parsed_data.json --output filled_checksheet.xlsx
    python populate_checksheet.py parsed_data.json --column cch  (fill CCH column)
    python populate_checksheet.py parsed_data.json --column source  (fill Source column)

Output:
    Creates filled checksheet Excel file.
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("ERROR: openpyxl not found. Install it:")
    print("  pip install openpyxl")
    sys.exit(1)


# =============================================================================
# CHECKSHEET CELL MAPPINGS
# =============================================================================

# Maps parsed data fields to checksheet cells
# Format: 'sheet_name': {'field_path': 'cell_address'}

CELL_MAPPINGS = {
    # 1099-INT tab - rows 5-14 for entries, row 25 for total
    '1099-INT': {
        'data_start_row': 5,
        'data_end_row': 14,
        'payer_col': 'B',
        'source_col': 'D',
        'cch_col': 'E',
        'total_row': 25,
    },
    # 1099-DIV tab
    '1099-DIV': {
        'data_start_row': 6,  # First payer section
        'rows_per_payer': 4,  # Header + name + ordinary + qualified
        'num_payers': 5,
        'payer_name_offset': 1,
        'ordinary_offset': 2,
        'qualified_offset': 3,
        'source_col': 'D',
        'cch_col': 'E',
        'total_row': 20,
    },
    # W-2 tab
    'W-2': {
        'data_start_row': 4,  # First employer section
        'rows_per_employer': 8,  # Section header + name + 6 boxes
        'num_employers': 3,
        'source_col': 'E',
        'cch_col': 'F',
        'total_wages_row': 30,
    },
    # 1099-R tab
    '1099-R': {
        'data_start_row': 5,
        'rows_per_payer': 5,
        'num_payers': 5,
        'source_col': 'E',
        'cch_col': 'F',
        'total_row': 20,
    },
    # SSA-1099 tab
    'SSA-1099': {
        'benefits_row': 5,
        'withholding_row': 6,
        'source_col': 'C',
        'cch_col': 'D',
    },
    # Schedule D tab
    'Schedule D': {
        'data_start_row': 5,
        'data_end_row': 14,
        'description_col': 'A',
        'term_col': 'B',
        'source_col': 'C',
        'cch_col': 'D',
        'total_row': 20,
    },
    # Sch C + E tab
    'Sch C + E': {
        'sch_c_start_row': 5,
        'sch_e_start_row': 20,
        'source_col': 'D',
        'cch_col': 'E',
    },
    # K-1 tab
    'K-1': {
        'partnership_start_row': 5,
        'scorp_start_row': 25,
        'source_col': 'D',
        'cch_col': 'E',
    },
    # NJ tab
    'NJ': {
        'data_start_row': 5,
        'source_col': 'C',
        'cch_col': 'D',
    },
    # Summary tab
    'Summary': {
        'client_name_cell': 'B3',
        'preparer_cell': 'B4',
        'federal_start_row': 8,
        'nj_start_row': 22,
        'source_col': 'B',
        'cch_col': 'C',
    },
}


def populate_1099int(ws, data: list, column: str = 'cch'):
    """Populate 1099-INT worksheet."""
    mapping = CELL_MAPPINGS['1099-INT']
    col = mapping['cch_col'] if column == 'cch' else mapping['source_col']
    
    row = mapping['data_start_row']
    for entry in data[:10]:  # Max 10 entries
        # Payer name
        payer = entry.get('payer_name', '')
        safe_write(ws, f"{mapping['payer_col']}{row}", payer)
        
        # Interest amount
        interest = entry.get('box1_interest', 0) or 0
        safe_write(ws, f"{col}{row}", int(round(interest)))
        
        row += 1
    
    return len(data)


def safe_write(ws, cell_ref, value):
    """Write to cell, skipping if it's a merged cell."""
    try:
        cell = ws[cell_ref]
        # Check if it's a merged cell
        if hasattr(cell, 'value'):
            cell.value = value
    except AttributeError:
        # Merged cell - skip it
        pass


def populate_1099div(ws, data: list, column: str = 'cch'):
    """Populate 1099-DIV worksheet."""
    mapping = CELL_MAPPINGS['1099-DIV']
    col = mapping['cch_col'] if column == 'cch' else mapping['source_col']
    
    row = mapping['data_start_row']
    for i, entry in enumerate(data[:5]):  # Max 5 payers
        if i > 0:
            row += mapping['rows_per_payer']
        
        # Payer name (row after section header)
        payer = entry.get('payer_name', '')
        safe_write(ws, f"B{row + mapping['payer_name_offset']}", payer)
        
        # Ordinary dividends
        ordinary = entry.get('box1a_ordinary_dividends', 0) or 0
        safe_write(ws, f"{col}{row + mapping['ordinary_offset']}", int(round(ordinary)))
        
        # Qualified dividends
        qualified = entry.get('box1b_qualified_dividends', 0) or 0
        safe_write(ws, f"{col}{row + mapping['qualified_offset']}", int(round(qualified)))
    
    return len(data)


def populate_w2(ws, data: list, column: str = 'cch'):
    """Populate W-2 worksheet."""
    mapping = CELL_MAPPINGS['W-2']
    col = mapping['cch_col'] if column == 'cch' else mapping['source_col']
    
    row = mapping['data_start_row']
    for i, entry in enumerate(data[:3]):  # Max 3 employers
        if i > 0:
            row += mapping['rows_per_employer']
        
        # Employer name (row 2 of section)
        employer = entry.get('employer_name', '')
        safe_write(ws, f"B{row + 1}", employer)
        
        # Box amounts start at row 3 of section
        box_row = row + 2
        
        wages = entry.get('box1_wages', 0) or 0
        safe_write(ws, f"{col}{box_row}", int(round(wages)))
        
        fed_wh = entry.get('box2_fed_withholding', 0) or 0
        safe_write(ws, f"{col}{box_row + 1}", int(round(fed_wh)))
        
        ss_wages = entry.get('box3_ss_wages', 0) or 0
        safe_write(ws, f"{col}{box_row + 2}", int(round(ss_wages)))
        
        medicare = entry.get('box5_medicare_wages', 0) or 0
        safe_write(ws, f"{col}{box_row + 3}", int(round(medicare)))
    
    return len(data)


def populate_1099r(ws, data: list, column: str = 'cch'):
    """Populate 1099-R worksheet."""
    mapping = CELL_MAPPINGS['1099-R']
    col = mapping['cch_col'] if column == 'cch' else mapping['source_col']
    
    row = mapping['data_start_row']
    for i, entry in enumerate(data[:5]):  # Max 5 payers
        if i > 0:
            row += mapping['rows_per_payer']
        
        # Payer name
        payer = entry.get('payer_name', '')
        safe_write(ws, f"B{row + 1}", payer)
        
        # Distribution code
        code = entry.get('box7_distribution_code', '')
        safe_write(ws, f"D{row + 1}", code)
        
        # Gross distribution (row + 2)
        gross = entry.get('box1_gross_distribution', 0) or 0
        safe_write(ws, f"{col}{row + 2}", int(round(gross)))
        
        # Taxable amount (row + 3)
        taxable = entry.get('box2a_taxable_amount', 0) or 0
        safe_write(ws, f"{col}{row + 3}", int(round(taxable)))
        
        # Fed withholding (row + 4)
        fed_wh = entry.get('box4_fed_withholding', 0) or 0
        safe_write(ws, f"{col}{row + 4}", int(round(fed_wh)))
    
    return len(data)


def populate_ssa1099(ws, data: list, column: str = 'cch'):
    """Populate SSA-1099 worksheet."""
    mapping = CELL_MAPPINGS['SSA-1099']
    col = mapping['cch_col'] if column == 'cch' else mapping['source_col']
    
    for entry in data[:1]:  # Usually just one
        benefits = entry.get('box5_net_benefits', 0) or 0
        safe_write(ws, f"{col}{mapping['benefits_row']}", int(round(benefits)))
        
        fed_wh = entry.get('box6_fed_withholding', 0) or 0
        safe_write(ws, f"{col}{mapping['withholding_row']}", int(round(fed_wh)))
    
    return len(data)


def populate_schedule_d(ws, data: list, column: str = 'cch'):
    """Populate Schedule D worksheet."""
    mapping = CELL_MAPPINGS['Schedule D']
    col = mapping['cch_col'] if column == 'cch' else mapping['source_col']
    
    row = mapping['data_start_row']
    for entry in data[:10]:  # Max 10 transactions
        # Description
        desc = entry.get('description', '')
        safe_write(ws, f"{mapping['description_col']}{row}", desc)
        
        # Term (ST/LT)
        term = entry.get('term', '')
        if term == 'S':
            term = 'ST'
        elif term == 'L':
            term = 'LT'
        safe_write(ws, f"{mapping['term_col']}{row}", term)
        
        # Gain/Loss (proceeds - cost)
        proceeds = entry.get('proceeds_actual', entry.get('proceeds', 0)) or 0
        cost = entry.get('cost_actual', entry.get('cost_basis', 0)) or 0
        gain = proceeds - cost
        safe_write(ws, f"{col}{row}", int(round(gain)))
        
        row += 1
    
    return len(data)


def populate_schedule_c(ws, data_c1: list, data_c2: list, column: str = 'cch'):
    """Populate Schedule C section of Sch C + E worksheet."""
    mapping = CELL_MAPPINGS['Sch C + E']
    col = mapping['cch_col'] if column == 'cch' else mapping['source_col']
    
    row = mapping['sch_c_start_row']
    for i, entry in enumerate(data_c1[:3]):  # Max 3 businesses
        if i > 0:
            row += 3  # Skip to next business section
        
        # Business name
        name = entry.get('business_name', '')
        safe_write(ws, f"A{row + 1}", name)
        
        # Gross receipts
        gross = entry.get('gross_receipts', 0) or 0
        safe_write(ws, f"{col}{row + 1}", int(round(gross)))
    
    return len(data_c1)


def populate_schedule_e(ws, data: list, column: str = 'cch'):
    """Populate Schedule E section of Sch C + E worksheet."""
    mapping = CELL_MAPPINGS['Sch C + E']
    col = mapping['cch_col'] if column == 'cch' else mapping['source_col']
    
    row = mapping['sch_e_start_row']
    for i, entry in enumerate(data[:5]):  # Max 5 properties
        if i > 0:
            row += 3  # Skip to next property section
        
        # Address
        address = entry.get('address', '')
        safe_write(ws, f"A{row + 1}", address)
        
        # Rents received
        rents = entry.get('line3_rents_received', 0) or 0
        safe_write(ws, f"{col}{row + 1}", int(round(rents)))
    
    return len(data)


def populate_k1(ws, partnership_data: list, scorp_data: list, trust_data: list, column: str = 'cch'):
    """Populate K-1 worksheet."""
    mapping = CELL_MAPPINGS['K-1']
    col = mapping['cch_col'] if column == 'cch' else mapping['source_col']
    
    # Partnership K-1s
    row = mapping['partnership_start_row']
    for entry in partnership_data[:3]:
        # Would need to determine which K-1 fields to map
        pass
    
    # S-Corp K-1s
    row = mapping['scorp_start_row']
    for entry in scorp_data[:2]:
        pass
    
    return len(partnership_data) + len(scorp_data)


def populate_checksheet(workbook, parsed_data: dict, column: str = 'cch'):
    """
    Populate the checksheet with parsed data.
    
    Args:
        workbook: openpyxl Workbook object
        parsed_data: dict from parse_input_listing.py
        column: 'cch' or 'source' - which column to fill
    """
    forms = parsed_data.get('forms', {})
    counts = {}
    
    # 1099-INT
    if 'IRS-1099INT' in forms and '1099-INT' in workbook.sheetnames:
        ws = workbook['1099-INT']
        counts['1099-INT'] = populate_1099int(ws, forms['IRS-1099INT'], column)
    
    # 1099-DIV
    if 'IRS-1099DIV' in forms and '1099-DIV' in workbook.sheetnames:
        ws = workbook['1099-DIV']
        counts['1099-DIV'] = populate_1099div(ws, forms['IRS-1099DIV'], column)
    
    # W-2
    if 'W-2' in forms and 'W-2' in workbook.sheetnames:
        ws = workbook['W-2']
        counts['W-2'] = populate_w2(ws, forms['W-2'], column)
    
    # 1099-R
    if 'IRS-1099R' in forms and '1099-R' in workbook.sheetnames:
        ws = workbook['1099-R']
        counts['1099-R'] = populate_1099r(ws, forms['IRS-1099R'], column)
    
    # SSA-1099
    if 'SSA-1099' in forms and 'SSA-1099' in workbook.sheetnames:
        ws = workbook['SSA-1099']
        counts['SSA-1099'] = populate_ssa1099(ws, forms['SSA-1099'], column)
    
    # Schedule D
    if 'D-1' in forms and 'Schedule D' in workbook.sheetnames:
        ws = workbook['Schedule D']
        counts['Schedule D'] = populate_schedule_d(ws, forms['D-1'], column)
    
    # Schedule C + E
    if 'Sch C + E' in workbook.sheetnames:
        ws = workbook['Sch C + E']
        if 'C-1' in forms:
            counts['Schedule C'] = populate_schedule_c(
                ws, forms.get('C-1', []), forms.get('C-2', []), column
            )
        if 'E-1' in forms:
            counts['Schedule E'] = populate_schedule_e(ws, forms['E-1'], column)
    
    return counts


def main():
    parser = argparse.ArgumentParser(
        description='Populate tax verification checksheet from parsed CCH data',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python populate_checksheet.py parsed_data.json
  python populate_checksheet.py parsed_data.json --checksheet blank_checksheet.xlsx
  python populate_checksheet.py parsed_data.json --output client_checksheet.xlsx
  python populate_checksheet.py parsed_data.json --column source  (fill Source column instead)

Workflow:
  1. Export Input Listing from CCH
  2. Run: python parse_input_listing.py input_listing.pdf
  3. Run: python populate_checksheet.py input_listing_parsed.json
  4. Review filled checksheet
        """
    )
    
    parser.add_argument(
        'json_file',
        help='Parsed data JSON file (from parse_input_listing.py)'
    )
    
    parser.add_argument(
        '--checksheet', '-c',
        help='Blank checksheet template (default: tax_verification_checksheet.xlsx in same dir)'
    )
    
    parser.add_argument(
        '--output', '-o',
        help='Output filled checksheet (default: <json_file>_checksheet.xlsx)'
    )
    
    parser.add_argument(
        '--column',
        choices=['cch', 'source'],
        default='cch',
        help='Which column to fill (default: cch)'
    )
    
    args = parser.parse_args()
    
    json_path = Path(args.json_file)
    
    if not json_path.exists():
        print(f"ERROR: File not found: {json_path}")
        sys.exit(1)
    
    # Load parsed data
    print(f"Loading: {json_path}")
    with open(json_path) as f:
        parsed_data = json.load(f)
    
    # Find checksheet template
    if args.checksheet:
        checksheet_path = Path(args.checksheet)
    else:
        # Look in same directory as script, then current directory
        script_dir = Path(__file__).parent
        possible_paths = [
            script_dir / 'tax_verification_checksheet.xlsx',
            Path('tax_verification_checksheet.xlsx'),
            Path.home() / 'tax_verification_checksheet.xlsx',
        ]
        checksheet_path = None
        for p in possible_paths:
            if p.exists():
                checksheet_path = p
                break
        
        if not checksheet_path:
            print("ERROR: Could not find checksheet template.")
            print("Specify with --checksheet or place tax_verification_checksheet.xlsx in:")
            for p in possible_paths:
                print(f"  {p}")
            sys.exit(1)
    
    if not checksheet_path.exists():
        print(f"ERROR: Checksheet not found: {checksheet_path}")
        sys.exit(1)
    
    print(f"Loading checksheet: {checksheet_path}")
    workbook = load_workbook(checksheet_path)
    
    # Populate
    print(f"Filling '{args.column}' column...")
    counts = populate_checksheet(workbook, parsed_data, args.column)
    
    # Determine output path
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = json_path.with_suffix('.xlsx')
        output_path = output_path.with_stem(json_path.stem.replace('_parsed', '') + '_checksheet')
    
    # Save
    workbook.save(output_path)
    
    print(f"\nOutput: {output_path}")
    
    # Print summary
    print("\n" + "="*60)
    print("POPULATION SUMMARY")
    print("="*60)
    print(f"  Column filled: {args.column.upper()}")
    print()
    for form_type, count in counts.items():
        print(f"  {form_type}: {count} entries")
    
    print("\n" + "="*60)
    print("NEXT STEPS")
    print("="*60)
    print("  1. Open the checksheet in Excel")
    print("  2. Fill in the Source column from source documents")
    print("  3. Review variances (Variance column should be 0)")
    print("  4. Check Summary tab for overall tie-out")


if __name__ == '__main__':
    main()
