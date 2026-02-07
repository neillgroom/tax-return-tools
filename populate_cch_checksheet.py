#!/usr/bin/env python3
# Python / Anaconda is located at: C:\Users\ngroom\AppData\Local\anaconda3\python.exe
"""
CCH 1040 Checksheet Populator
=============================
Populates the CCH_1040_Checksheet_current.xlsx from parsed source documents.

This script is specifically designed for the multi-column checksheet layout:
- Income sheet: Columns B-F for up to 5 sources, Column H for "On 1040" (CCH)
- Source data goes into B-F columns
- CCH/return data goes into Column H

Usage:
    python populate_cch_checksheet.py parsed_data.json
    python populate_cch_checksheet.py parsed_data.json --checksheet "C:\\Tax\\CCH_1040_Checksheet.xlsx"
    python populate_cch_checksheet.py parsed_data.json --mode source  (fill source columns B-F)
    python populate_cch_checksheet.py parsed_data.json --mode cch     (fill CCH column H)
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime
import os

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("ERROR: openpyxl not found. Install with: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# CELL COLOR DEFINITIONS (for quality tracking)
# =============================================================================

# Pale yellow - Manual entry items (can't upload to CCH)
FILL_MANUAL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

# Pale orange - OCR data (not 100% accurate, needs verification)
FILL_OCR = PatternFill(start_color="FFE4CC", end_color="FFE4CC", fill_type="solid")

# Pale red - Validation issues or suspicious values
FILL_ISSUE = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

# Pale blue - Partial extraction (some fields missing)
FILL_PARTIAL = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

# Forms that cannot be imported to CCH (manual entry required)
MANUAL_ENTRY_FORMS = {'PROPERTY-TAX', 'K-1'}  # Add more as identified


# =============================================================================
# INCOME SHEET MAPPINGS (CCH_1040_Checksheet_current.xlsx)
# =============================================================================

INCOME_SHEET = {
    'name': 'Income',

    # W-2 section (rows 5-15, shifted after adding Box 3 and Box 5)
    'W-2': {
        'header_row': 5,
        'data_cols': ['B', 'C', 'D', 'E', 'F'],  # Up to 5 employers
        'cch_col': 'H',  # "On 1040" column
        'rows': {
            'box1_wages': 6,
            'box2_fed_withholding': 7,
            'box3_ss_wages': 8,
            'box5_medicare_wages': 9,
            'box4_ss_tax': 10,
            'box6_medicare_tax': 11,
            'box12_retirement': 12,
            'box16_state_wages': 13,
            'box17_state_withholding': 14,
        },
        'name_row': 5,  # Where employer names go (header row)
    },

    # 1099-INT section (shifted +2 rows)
    '1099-INT': {
        'header_row': 18,
        'data_cols': ['B', 'C', 'D', 'E', 'F'],
        'cch_col': 'H',
        'rows': {
            'box1_interest': 19,
            'box3_savings_bond': 20,
            'box8_tax_exempt': 21,
            'box4_fed_withholding': 22,
        },
        'name_row': 18,
    },

    # 1099-DIV section (shifted +2 rows)
    '1099-DIV': {
        'header_row': 26,
        'data_cols': ['B', 'C', 'D', 'E', 'F'],
        'cch_col': 'H',
        'rows': {
            'box1a_ordinary_dividends': 27,
            'box1b_qualified_dividends': 28,
            'box2a_total_cap_gain': 29,
            'box3_nondiv_dist': 30,
            'box5_sec199a': 31,
            'box7_foreign_tax': 32,
            'box12_exempt_int_div': 33,
            'box4_fed_withholding': 34,
        },
        'name_row': 26,
    },

    # 1099-R section (shifted +2 rows)
    '1099-R': {
        'header_row': 38,
        'data_cols': ['B', 'C', 'D', 'E', 'F'],
        'cch_col': 'H',
        'rows': {
            'box1_gross_distribution': 39,
            'box2a_taxable_amount': 40,
            'box4_fed_withholding': 41,
            'box14_state_withholding': 42,
            'box7_distribution_code': 43,
        },
        'name_row': 38,
    },

    # SSA-1099 section (shifted +2 rows)
    'SSA-1099': {
        'header_row': 47,
        'data_cols': ['B', 'C', 'D', 'E', 'F'],
        'cch_col': 'H',
        'rows': {
            'box3_benefits_paid': 48,
            'box4_benefits_repaid': 49,
            'box5_net_benefits': 50,
            'box6_fed_withholding': 51,
        },
        'name_row': 47,
    },
}

# =============================================================================
# K-1 SHEET MAPPINGS
# =============================================================================

K1_SHEET = {
    'name': 'K-1s',
    'entity_cols': ['C', 'D', 'E'],  # Up to 3 entities
    'entity_name_row': 4,  # Row for entity names in header
    'rows': {
        'box1_ordinary_income': 5,      # Line 1 - Ordinary business income
        'box2_net_rental_income': 6,    # Line 2 - Net rental real estate
        'box3_other_net_rental_income': 7,  # Line 3 - Other net rental
        'box4_guaranteed_payments': 8,  # Line 4a - Guaranteed payments for services
        # 'box4b_guaranteed_capital': 9,  # Line 4b - Guaranteed payments for capital
        # 'box4c_total_guaranteed': 10,   # Line 4c - Total guaranteed payments
        'box5_interest_income': 11,     # Line 5 - Interest income
        'box6a_ordinary_dividends': 12, # Line 6a - Ordinary dividends
        'box6b_qualified_dividends': 13, # Line 6b - Qualified dividends
        # 'box6c_dividend_equivalents': 14, # Line 6c - Dividend equivalents
        'box7_royalties': 15,           # Line 7 - Royalties
        'box8_net_st_cap_gain': 16,     # Line 8 - Net short-term capital gain
        'box9a_net_lt_cap_gain': 17,    # Line 9a - Net long-term capital gain
        # 'box9b_collectibles': 18,       # Line 9b - Collectibles (28%) gain
        # 'box9c_unrecap_1250': 19,       # Line 9c - Unrecaptured section 1250
        'box10_net_1231_gain': 20,      # Line 10 - Net section 1231 gain
        'box11_other_income': 21,       # Line 11 - Other income
        'box12_sec179_deduction': 22,   # Line 12 - Section 179 deduction
        'box13_other_deductions': 23,   # Line 13 - Other deductions
        'box14_self_employment': 24,    # Line 14 - Self-employment earnings
    }
}


# =============================================================================
# SCHEDULE A MAPPINGS (Itemized Deductions)
# =============================================================================

SCHEDULE_A_SHEET = {
    'name': 'Schedule A',

    # 1098 Mortgage Interest goes to row 16 (8a - Home mortgage interest)
    '1098': {
        'rows': {
            'box1_mortgage_interest': 16,  # 8a - Home mortgage interest
            'box5_mortgage_insurance': 19,  # 8d - Mortgage insurance premiums
        },
        'source_col': 'E',  # Source Amount column
        'entered_col': 'F',  # Entered Amount column
    },

    # Property Tax goes to row 10 (5b - State/local real estate taxes)
    'PROPERTY-TAX': {
        'rows': {
            'ad_valorem_taxes': 10,  # 5b - State/local real estate taxes
        },
        'source_col': 'E',
        'entered_col': 'F',
    },
}


def safe_write(ws, cell_ref, value, fill=None):
    """Write to cell safely, handling merged cells. Optionally apply fill color."""
    try:
        cell = ws[cell_ref]
        if hasattr(cell, 'value'):
            cell.value = value
            if fill is not None:
                cell.fill = fill
    except (AttributeError, KeyError):
        pass


def get_quality_fill(entry, form_type=None):
    """
    Determine the appropriate cell fill color based on quality metadata.

    Priority order:
    1. FILL_ISSUE (red) - Validation issues, math errors, or very low confidence
    2. FILL_PARTIAL (blue) - Missing required fields
    3. FILL_MANUAL (yellow) - Can't upload to CCH (like property tax)
    4. FILL_OCR (orange) - OCR sourced data
    5. None - Clean digital data
    """
    quality = entry.get('_quality', {})

    # Check for validation issues (highest priority)
    if quality.get('issues') or quality.get('math_errors'):
        return FILL_ISSUE

    # Check for very low confidence (below 50%)
    if quality.get('overall_confidence', 100) < 50:
        return FILL_ISSUE

    # Check for missing required fields
    if quality.get('missing_required'):
        return FILL_PARTIAL

    # Check if this is a manual entry form type
    if form_type and form_type in MANUAL_ENTRY_FORMS:
        return FILL_MANUAL

    # Check if data came from OCR
    if quality.get('is_ocr', False):
        return FILL_OCR

    # Clean digital data - no fill
    return None


def clear_income_sheet(ws):
    """Clear all data from Income sheet columns B-F (source data columns)."""
    sections = [
        # W-2: rows 5-14 (shifted after adding Box 3 and Box 5)
        {'name_row': 5, 'data_rows': range(6, 15)},
        # 1099-INT: rows 18-22 (shifted +2)
        {'name_row': 18, 'data_rows': range(19, 23)},
        # 1099-DIV: rows 26-34 (shifted +2)
        {'name_row': 26, 'data_rows': range(27, 35)},
        # 1099-R: rows 38-43 (shifted +2)
        {'name_row': 38, 'data_rows': range(39, 44)},
        # SSA-1099: rows 47-51 (shifted +2)
        {'name_row': 47, 'data_rows': range(48, 52)},
    ]

    for section in sections:
        # Clear header row names (B-F)
        for col in ['B', 'C', 'D', 'E', 'F']:
            safe_write(ws, f"{col}{section['name_row']}", None)
        # Clear data rows (B-F)
        for row in section['data_rows']:
            for col in ['B', 'C', 'D', 'E', 'F']:
                safe_write(ws, f"{col}{row}", None)


def update_audit_sheet(workbook, parsed_data, mode):
    """Update the Audit sheet with extraction log."""
    if 'Audit' not in workbook.sheetnames:
        return

    ws = workbook['Audit']

    # Find or set client name from metadata
    folder = parsed_data.get('metadata', {}).get('source_folder', 'Unknown')
    client_name = Path(folder).parent.name if folder else 'Unknown'

    # Update header info
    safe_write(ws, 'A1', f"Client: {client_name}")
    safe_write(ws, 'A2', f"Processed: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    safe_write(ws, 'A3', f"Mode: {mode.upper()}")

    # Color Legend
    ws['E1'] = "COLOR LEGEND:"
    ws['E1'].font = Font(bold=True)
    safe_write(ws, 'E2', "Manual Entry", FILL_MANUAL)
    safe_write(ws, 'F2', "Can't upload to CCH - manual entry required")
    safe_write(ws, 'E3', "OCR Data", FILL_OCR)
    safe_write(ws, 'F3', "Extracted via OCR - verify accuracy")
    safe_write(ws, 'E4', "Issue Found", FILL_ISSUE)
    safe_write(ws, 'F4', "Validation issue or low confidence")
    safe_write(ws, 'E5', "Partial Data", FILL_PARTIAL)
    safe_write(ws, 'F5', "Some required fields missing")

    # Find starting row for extraction log (after any existing content)
    log_start = 5

    safe_write(ws, 'A' + str(log_start), "EXTRACTION LOG")
    safe_write(ws, 'A' + str(log_start + 1), "Form Type")
    safe_write(ws, 'B' + str(log_start + 1), "Source File")
    safe_write(ws, 'C' + str(log_start + 1), "Key Value")
    safe_write(ws, 'D' + str(log_start + 1), "Payer/Employer")

    row = log_start + 2
    forms = parsed_data.get('forms', {})

    for form_type, entries in forms.items():
        for entry in entries:
            safe_write(ws, f'A{row}', form_type)
            safe_write(ws, f'B{row}', entry.get('source_file', ''))

            # Get key value based on form type
            if form_type == 'W-2':
                key_val = entry.get('box1_wages', 0)
                payer = entry.get('employer_name', '')
            elif form_type in ['IRS-1099INT', '1099-INT']:
                key_val = entry.get('box1_interest', 0)
                payer = entry.get('payer_name', '')
            elif form_type in ['IRS-1099DIV', '1099-DIV']:
                key_val = entry.get('box1a_ordinary_dividends', 0)
                payer = entry.get('payer_name', '')
            elif form_type in ['IRS-1099R', '1099-R']:
                key_val = entry.get('box1_gross_distribution', 0)
                payer = entry.get('payer_name', '')
            elif form_type == 'SSA-1099':
                key_val = entry.get('box5_net_benefits', 0)
                payer = entry.get('description', 'Social Security')
            elif form_type == '1098':
                key_val = entry.get('box1_mortgage_interest', 0)
                payer = entry.get('lender_name', '')
            elif form_type == 'PROPERTY-TAX':
                key_val = entry.get('ad_valorem_taxes', 0)
                payer = entry.get('county', '')
            else:
                key_val = ''
                payer = ''

            safe_write(ws, f'C{row}', f"${key_val:,.2f}" if isinstance(key_val, (int, float)) else key_val)
            safe_write(ws, f'D{row}', payer)
            row += 1

    # Summary
    row += 1
    safe_write(ws, f'A{row}', f"Total forms extracted: {sum(len(e) for e in forms.values())}")


def format_amount(value):
    """Format a numeric value for the checksheet."""
    if value is None or value == '':
        return None
    try:
        num = float(str(value).replace(',', '').replace('$', ''))
        if num == 0:
            return None
        return round(num)
    except (ValueError, TypeError):
        return None


# Minimum confidence threshold for populating values
MIN_CONFIDENCE_THRESHOLD = 60


def should_skip_entry(item):
    """
    Check if an entire form entry should be skipped due to quality issues.
    Returns (skip: bool, reason: str or None)
    """
    quality = item.get('_quality', {})

    # Skip if overall confidence is too low
    overall = quality.get('overall_confidence', 100)
    if overall < MIN_CONFIDENCE_THRESHOLD:
        return True, f"Low confidence ({overall}%)"

    # Skip if there are math errors (indicates wrong values parsed)
    math_errors = quality.get('math_errors', [])
    if math_errors:
        return True, f"Math validation failed ({len(math_errors)} errors)"

    # Skip if payer/employer name looks like garbage (form labels, etc.)
    name = item.get('employer_name', '') or item.get('payer_name', '') or ''
    name_lower = name.lower()
    garbage_patterns = ['zip', '1099', 'box ', 'federal', 'withheld', 'income tax',
                        'postal code', 'telephone', 'payer', 'recipient', 'form w-2',
                        'fed.', 'medicare', 'social security', 'wages', 'w-2 box']
    for pattern in garbage_patterns:
        if pattern in name_lower:
            return True, f"Invalid name (contains '{pattern}')"

    # Skip if interest amount looks like a tax year
    interest = item.get('box1_interest', 0)
    if interest in [2023, 2024, 2025, 2026]:
        return True, f"Interest amount looks like year ({int(interest)})"

    return False, None


def should_skip_field(item, field_name):
    """
    Check if a specific field should be skipped due to quality issues.
    Returns True if field should be skipped.
    """
    quality = item.get('_quality', {})
    field_quality = quality.get('field_quality', {})

    # Check if this specific field has low confidence
    if field_name in field_quality:
        fq = field_quality[field_name]
        confidence = fq.get('confidence', 100)
        if confidence < MIN_CONFIDENCE_THRESHOLD:
            return True
        # Skip if field has issues flagged
        if fq.get('issues'):
            return True

    return False


def populate_w2_source(ws, w2_data, mapping):
    """Populate W-2 data into source columns."""
    count = 0
    skipped = 0
    for idx, w2 in enumerate(w2_data[:5]):  # Max 5 employers
        col = mapping['data_cols'][idx]

        # Check if entire entry should be skipped
        skip, reason = should_skip_entry(w2)
        if skip:
            skipped += 1
            print(f"    Skipping W-2 '{w2.get('employer_name', 'Unknown')[:30]}': {reason}")
            continue

        # Determine fill color based on quality
        fill = get_quality_fill(w2, 'W-2')

        # Write employer name in header row
        employer = w2.get('employer_name', f'Employer {idx+1}')
        safe_write(ws, f"{col}{mapping['name_row']}", employer, fill)

        # Write box values (skip fields with quality issues)
        for field, row in mapping['rows'].items():
            if should_skip_field(w2, field):
                continue
            value = format_amount(w2.get(field, 0))
            if value:
                safe_write(ws, f"{col}{row}", value, fill)

        count += 1
    if skipped:
        print(f"    ({skipped} W-2(s) skipped due to quality issues)")
    return count


def populate_1099_source(ws, data_list, mapping, form_type):
    """Populate 1099 data (INT, DIV, R) into source columns."""
    count = 0
    skipped = 0
    for idx, item in enumerate(data_list[:5]):  # Max 5 payers
        col = mapping['data_cols'][idx]

        # Check if entire entry should be skipped
        skip, reason = should_skip_entry(item)
        if skip:
            skipped += 1
            print(f"    Skipping {form_type} '{item.get('payer_name', 'Unknown')[:30]}': {reason}")
            continue

        # Determine fill color based on quality
        fill = get_quality_fill(item, form_type)

        # Write payer name in header row
        payer = item.get('payer_name', f'Payer {idx+1}')
        safe_write(ws, f"{col}{mapping['name_row']}", payer, fill)

        # Write box values (skip fields with quality issues)
        for field, row in mapping['rows'].items():
            if should_skip_field(item, field):
                continue
            value = format_amount(item.get(field, 0))
            if value:
                safe_write(ws, f"{col}{row}", value, fill)

        count += 1
    if skipped:
        print(f"    ({skipped} {form_type}(s) skipped due to quality issues)")
    return count


def populate_ssa_source(ws, ssa_data, mapping):
    """Populate SSA-1099 data into source columns."""
    count = 0
    skipped = 0
    for idx, item in enumerate(ssa_data[:5]):
        col = mapping['data_cols'][idx]

        # Check if entire entry should be skipped
        skip, reason = should_skip_entry(item)
        if skip:
            skipped += 1
            print(f"    Skipping SSA-1099: {reason}")
            continue

        # Determine fill color based on quality
        fill = get_quality_fill(item, 'SSA-1099')

        # SSA typically just has one entry per person
        name = item.get('description', 'Social Security')
        safe_write(ws, f"{col}{mapping['name_row']}", name, fill)

        for field, row in mapping['rows'].items():
            if should_skip_field(item, field):
                continue
            value = format_amount(item.get(field, 0))
            if value:
                safe_write(ws, f"{col}{row}", value, fill)

        count += 1
    if skipped:
        print(f"    ({skipped} SSA-1099(s) skipped due to quality issues)")
    return count


def populate_cch_column(ws, totals, form_type, mapping):
    """Populate the CCH/On 1040 column (H) with totals or CCH data."""
    col = mapping['cch_col']

    for field, row in mapping['rows'].items():
        value = format_amount(totals.get(field, 0))
        if value:
            safe_write(ws, f"{col}{row}", value)


def populate_schedule_a(workbook, parsed_data, mode='source'):
    """Populate Schedule A (Itemized Deductions) with 1098 and property tax data."""
    counts = {}
    forms = parsed_data.get('forms', {})

    if SCHEDULE_A_SHEET['name'] not in workbook.sheetnames:
        print(f"  WARNING: '{SCHEDULE_A_SHEET['name']}' sheet not found")
        return counts

    ws = workbook[SCHEDULE_A_SHEET['name']]

    # 1098 Mortgage Interest
    if '1098' in forms and forms['1098']:
        mapping = SCHEDULE_A_SHEET['1098']
        col = mapping['source_col'] if mode == 'source' else mapping['entered_col']

        # Determine fill color - check if any 1098 has OCR or issues
        fill = None
        for entry in forms['1098']:
            entry_fill = get_quality_fill(entry, '1098')
            if entry_fill:
                fill = entry_fill
                break

        # Sum all 1098 entries (if multiple mortgages)
        total_interest = sum(f.get('box1_mortgage_interest', 0) for f in forms['1098'])
        total_insurance = sum(f.get('box5_mortgage_insurance', 0) for f in forms['1098'])

        if total_interest > 0:
            safe_write(ws, f"{col}{mapping['rows']['box1_mortgage_interest']}", round(total_interest), fill)
            counts['1098'] = len(forms['1098'])
            fill_msg = " [OCR]" if fill == FILL_OCR else " [ISSUE]" if fill == FILL_ISSUE else ""
            print(f"  1098 Mortgage Interest: ${total_interest:,.2f}{fill_msg}")

        if total_insurance > 0:
            safe_write(ws, f"{col}{mapping['rows']['box5_mortgage_insurance']}", round(total_insurance), fill)

    # Property Tax - ALWAYS yellow (manual entry - can't upload to CCH)
    if 'PROPERTY-TAX' in forms and forms['PROPERTY-TAX']:
        mapping = SCHEDULE_A_SHEET['PROPERTY-TAX']
        col = mapping['source_col'] if mode == 'source' else mapping['entered_col']

        # Property tax is always manual entry (yellow), but check for OCR/issues too
        fill = FILL_MANUAL  # Base color for manual entry
        for entry in forms['PROPERTY-TAX']:
            quality = entry.get('_quality', {})
            if quality.get('issues') or quality.get('math_errors'):
                fill = FILL_ISSUE  # Issues take priority
                break
            if quality.get('is_ocr'):
                fill = FILL_OCR  # OCR takes priority over manual

        # Sum all property tax entries
        total_property_tax = sum(f.get('ad_valorem_taxes', 0) for f in forms['PROPERTY-TAX'])

        if total_property_tax > 0:
            safe_write(ws, f"{col}{mapping['rows']['ad_valorem_taxes']}", round(total_property_tax), fill)
            counts['PROPERTY-TAX'] = len(forms['PROPERTY-TAX'])
            fill_msg = " [MANUAL ENTRY]" if fill == FILL_MANUAL else " [OCR]" if fill == FILL_OCR else " [ISSUE]" if fill == FILL_ISSUE else ""
            print(f"  Property Tax: ${total_property_tax:,.2f}{fill_msg}")

    return counts


def populate_k1_sheet(workbook, parsed_data, mode='source'):
    """Populate K-1s sheet with K-1 data."""
    counts = {}
    forms = parsed_data.get('forms', {})

    if K1_SHEET['name'] not in workbook.sheetnames:
        print(f"  WARNING: '{K1_SHEET['name']}' sheet not found")
        return counts

    ws = workbook[K1_SHEET['name']]

    k1_data = forms.get('K-1', [])
    if not k1_data:
        return counts

    entity_cols = K1_SHEET['entity_cols']

    for idx, entry in enumerate(k1_data):
        if idx >= len(entity_cols):
            print(f"  WARNING: More than {len(entity_cols)} K-1s, skipping extras")
            break

        col = entity_cols[idx]
        entity_name = entry.get('entity_name', f'Entity {idx+1}')

        # Determine fill color based on quality
        fill = get_quality_fill(entry, 'K-1')

        # Write entity name in header row
        safe_write(ws, f"{col}{K1_SHEET['entity_name_row']}", entity_name[:25], fill)

        # Write each K-1 line item
        for field, row in K1_SHEET['rows'].items():
            value = entry.get(field, 0)
            if value and value != 0:
                safe_write(ws, f"{col}{row}", value, fill)

    counts['K-1'] = len(k1_data)
    fill_msg = " [MANUAL ENTRY]"  # K-1s always need manual entry
    print(f"  K-1: {len(k1_data)} entries{fill_msg}")

    return counts


def populate_checksheet(workbook, parsed_data, mode='source'):
    """
    Populate the checksheet from parsed data.

    mode='source': Fill columns B-F with source document data
    mode='cch': Fill column H with CCH/return data
    """
    counts = {}
    forms = parsed_data.get('forms', {})

    # Get Income sheet
    if INCOME_SHEET['name'] not in workbook.sheetnames:
        print(f"WARNING: '{INCOME_SHEET['name']}' sheet not found")
        return counts

    ws = workbook[INCOME_SHEET['name']]

    # Clear existing data first
    if mode == 'source':
        print("  Clearing existing source data...")
        clear_income_sheet(ws)

    # Update Audit sheet
    print("  Updating Audit sheet...")
    update_audit_sheet(workbook, parsed_data, mode)

    # W-2
    if 'W-2' in forms and forms['W-2']:
        mapping = INCOME_SHEET['W-2']
        if mode == 'source':
            counts['W-2'] = populate_w2_source(ws, forms['W-2'], mapping)
        else:
            # For CCH mode, we'd need CCH totals - not source docs
            pass
        print(f"  W-2: {len(forms['W-2'])} entries")

    # 1099-INT
    int_key = 'IRS-1099INT' if 'IRS-1099INT' in forms else '1099-INT'
    if int_key in forms and forms[int_key]:
        mapping = INCOME_SHEET['1099-INT']
        if mode == 'source':
            counts['1099-INT'] = populate_1099_source(ws, forms[int_key], mapping, '1099-INT')
        print(f"  1099-INT: {len(forms[int_key])} entries")

    # 1099-DIV
    div_key = 'IRS-1099DIV' if 'IRS-1099DIV' in forms else '1099-DIV'
    if div_key in forms and forms[div_key]:
        mapping = INCOME_SHEET['1099-DIV']
        if mode == 'source':
            counts['1099-DIV'] = populate_1099_source(ws, forms[div_key], mapping, '1099-DIV')
        print(f"  1099-DIV: {len(forms[div_key])} entries")

    # 1099-R
    r_key = 'IRS-1099R' if 'IRS-1099R' in forms else '1099-R'
    if r_key in forms and forms[r_key]:
        mapping = INCOME_SHEET['1099-R']
        if mode == 'source':
            counts['1099-R'] = populate_1099_source(ws, forms[r_key], mapping, '1099-R')
        print(f"  1099-R: {len(forms[r_key])} entries")

    # SSA-1099
    if 'SSA-1099' in forms and forms['SSA-1099']:
        mapping = INCOME_SHEET['SSA-1099']
        if mode == 'source':
            counts['SSA-1099'] = populate_ssa_source(ws, forms['SSA-1099'], mapping)
        print(f"  SSA-1099: {len(forms['SSA-1099'])} entries")

    # Populate Schedule A (Itemized Deductions) for 1098 and property tax
    print("  Populating Schedule A...")
    sched_a_counts = populate_schedule_a(workbook, parsed_data, mode)
    counts.update(sched_a_counts)

    # Populate K-1s sheet
    print("  Populating K-1s...")
    k1_counts = populate_k1_sheet(workbook, parsed_data, mode)
    counts.update(k1_counts)

    return counts


def main():
    parser = argparse.ArgumentParser(
        description='Populate CCH 1040 Checksheet from parsed tax documents',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python populate_cch_checksheet.py parsed_data.json
  python populate_cch_checksheet.py parsed_data.json --checksheet "C:\\Tax\\Checksheet.xlsx"
  python populate_cch_checksheet.py parsed_data.json --mode source
  python populate_cch_checksheet.py parsed_data.json --mode cch

Modes:
  source - Fill columns B-F with source document data (default)
  cch    - Fill column H with CCH/return data
        """
    )

    parser.add_argument('json_file', help='Parsed data JSON file')
    parser.add_argument('--checksheet', '-c',
                        default=r'C:\Tax\CCH_1040_Checksheet_current.xlsx',
                        help='Checksheet template file')
    parser.add_argument('--output', '-o', help='Output file (default: adds _filled to input)')
    parser.add_argument('--mode', '-m', choices=['source', 'cch'], default='source',
                        help='Which columns to fill (default: source)')

    args = parser.parse_args()

    json_path = Path(args.json_file)
    if not json_path.exists():
        print(f"ERROR: File not found: {json_path}")
        sys.exit(1)

    checksheet_path = Path(args.checksheet)
    if not checksheet_path.exists():
        print(f"ERROR: Checksheet not found: {checksheet_path}")
        sys.exit(1)

    # Load data
    print(f"Loading: {json_path}")
    with open(json_path) as f:
        parsed_data = json.load(f)

    # Load checksheet
    print(f"Loading checksheet: {checksheet_path}")
    workbook = load_workbook(checksheet_path)

    # Populate
    print(f"\nPopulating {args.mode.upper()} columns...")
    counts = populate_checksheet(workbook, parsed_data, args.mode)

    # Determine output path - save as client_name_year_checksheet.xlsx in client folder
    if args.output:
        output_path = Path(args.output)
    else:
        # Get client name and year from folder structure
        # JSON is in L:\ClientName\2025\2025_parsed.json
        client_folder = json_path.parent  # L:\ClientName\2025
        year = client_folder.name  # 2025
        client_name = client_folder.parent.name  # ClientName

        # Clean up client name for filename (replace spaces with underscores)
        client_name_clean = client_name.replace(' ', '_')

        output_filename = f"{client_name_clean}_{year}_checksheet.xlsx"
        output_path = client_folder / output_filename

    # Save
    workbook.save(output_path)

    print(f"\n{'='*60}")
    print("POPULATION COMPLETE")
    print('='*60)
    print(f"Mode: {args.mode.upper()}")
    print(f"Output: {output_path}")
    print()
    print("Forms populated:")
    for form_type, count in counts.items():
        print(f"  {form_type}: {count}")


if __name__ == '__main__':
    main()
