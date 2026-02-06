#!/usr/bin/env python3
"""
CCH Import Excel Generator
==========================
Generates Excel files compatible with CCH ProSystem fx worksheet import.

This script reads the JSON output from parse_source_docs.py and generates
Excel files that can be directly imported into CCH worksheets.

Usage:
    python generate_cch_import.py parsed_data.json
    python generate_cch_import.py parsed_data.json --output-dir "L:\\Client\\2025"
    python generate_cch_import.py parsed_data.json --forms INT,DIV,SSA

Supported forms:
    W2   - W-2 (Wages and Salaries)
    INT  - 1099-INT (Interest)
    DIV  - 1099-DIV (Dividends)
    SSA  - SSA-1099 (Social Security)
    MTG  - 1098 (Mortgage Interest)
    CAP  - 1099-B (Capital Gains) [future]
    K1P  - K-1 Partnership [future]
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("ERROR: openpyxl not found. Install with: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# CCH COLUMN DEFINITIONS (based on actual CCH exports)
# =============================================================================

# 1099-INT: Interest Worksheet (83 columns)
# Headers are in Row 6, Data starts Row 7
INT_COLUMNS = [
    "TSJ", "Payer", "Interest Income", "U.S. Savings Bonds", "Prior Year",
    "Federal Tax Withheld", "Investment Expenses", "Early Withdrawal Penalty",
    "Tax-Exempt Interest", "Tax-Exempt Default", "State NonTaxable - Override",
    "State Taxable - Override", "Code", "Amount or % Subject to AMT - Override",
    "Foreign Tax Paid", "Foreign Country Code", "Foreign Income Category",
    "Foreign Country Name - Override",
    "Re-Source US Income to Foreign Category Specified in Tax Treaty",
    "FS", "State", "City ", "Dual Status", "8938 Category",
    "Foreign Taxes Accrued", "Foreign Taxes Paid Date", "Street address",
    "City", "State", "ZIP", "Foreign Postal Code", "Foreign Country Code",
    "Foreign Province / State / County", "EIN", "Phone No.", "RTN",
    "Specified Private Activity Bond Interest", "CUSIP No.", "Market Discount",
    "Bond Premium", "Bond Premium on Treasury Obligations",
    "Bond Premium on Tax-Exempt Bond", "OID Income", "State Distribution Amount",
    "Other Periodic Interest", "Tax-Exempt OID", "Branch Reporting Elected",
    "Special Interest Amount", "Special Interest Type", "State Tax Withheld",
    "State Identification No.", "State Percent", "State Amount", "State Use",
    "New Hampshire Nontaxable Amount", "New Hampshire Nontaxable Reason Code",
    "Illinois Nontaxable Amount", "Illinois Nontaxable Reason Code",
    "Account Number", "2nd TIN Not.", "Pro Forma / Organizer Option",
    "OID Description", "OID Discount on Treasury Bonds", "FATCA Filing Requirement",
    "Foreign Amount (If Different)", "Exclude Record from Schedule B",
    "Conversion Rate", "Hypothetical State Code (If Different)",
    "Exclude from Hypothetical", "Home Buyer's SSN", "Street Address", "City",
    "State", "ZIP / Postal Code", "Foreign Country Code", "Name", "SSN",
    "Street Address", "City", "State", "ZIP / Postal Code", "Foreign Country Code",
    "Foreign Province / State / County"
]

# Key column positions (1-indexed) for 1099-INT
INT_MAP = {
    'tsj': 1,
    'payer': 2,
    'interest_income': 3,      # Box 1
    'us_savings_bonds': 4,     # Box 3
    'fed_withholding': 6,      # Box 4
    'tax_exempt_interest': 9,  # Box 8
    'ein': 34,
}

# 1099-DIV: Dividends Worksheet (75 columns)
DIV_COLUMNS = [
    "TSJ", "Payer Name", "Ordinary Dividends", "Prior Year", "Qualified Dividends",
    "Total Capital Gain Distribution", "Unrecaptured Section 1250 Gain",
    "Section 1202 Gain", "Collectibles (28%) Gain", "Nondividend Distributions",
    "Federal Tax Withheld", "Investment Expenses", "Exempt-Interest Dividends",
    "Tax-Exempt Default", "State Nontaxable - Override", "State Taxable - Override",
    "Tax Exempt Interest Code", "Amount or % Subject to AMT - Override",
    "Foreign Tax Paid", "Foreign Country Code", "Foreign Income Category",
    "Foreign Country Name - Override",
    "Re-Source US Income to Foreign Category Specified in Tax Treaty",
    "U.S. Bonds - Amount or % in Box 1 (Ordinary Div.)",
    "Cash Liquidation Distribution", "Noncash Liquidation Distribution",
    "FS", "State", "City", "Dual Status",
    "Foreign Ordinary Dividends (If Different)",
    "Foreign Qualified Dividends (If Different)",
    "Foreign Capital Gain Distribution", "Exclude Record from Schedule B",
    "8938 Category", "Foreign Taxes Accrued", "Foreign Taxes Date Paid",
    "Street Address", "City", "State", "ZIP", "Foreign Postal Code",
    "Foreign Country Code", "Foreign Province / State / County", "EIN",
    "Phone No.", "Specified Private Activity Bond Interest Dividends",
    "State Tax Withheld", "State Identification No.", "State Percent",
    "State Amount", "State Use", "New Hampshire Nontaxable Amount",
    "New Hampshire Nontaxable Reason Code", "Account No.", "2nd TIN Not.",
    "Pro Forma / Organizer Option", "Nominee Dividends", "FATCA Filing Requirement",
    "Foreign Unrecaptured 1250 Gain (If Different)",
    "Foreign 1202 Gain (If Different)",
    "Foreign Collectibles (28%) Gain (If Different)",
    "Taxes Paid (Foreign Amount)", "Conversion Rate", "Exclude from Hypothetical",
    "Hypothetical State Code (If Different)", "Name", "Social Security No.",
    "Street Address", "City", "State", "ZIP / Postal Code", "Foreign Country Code",
    "Foreign Province / State / County", "Section 199A Dividends"
]

# Key column positions (1-indexed) for 1099-DIV
DIV_MAP = {
    'tsj': 1,
    'payer_name': 2,
    'ordinary_dividends': 3,         # Box 1a
    'qualified_dividends': 5,        # Box 1b
    'total_cap_gain_dist': 6,        # Box 2a
    'unrecap_1250': 7,               # Box 2b
    'section_1202': 8,               # Box 2c
    'collectibles_28': 9,            # Box 2d
    'nondividend_dist': 10,          # Box 3
    'fed_withholding': 11,           # Box 4
    'exempt_int_dividends': 13,      # Box 12
    'foreign_tax_paid': 19,          # Box 7
    'ein': 45,
    'section_199a_dividends': 75,    # Box 5
}

# SSA-1099: Social Security Benefit Statement (35 columns)
SSA_COLUMNS = [
    "TSJ", "Name", "Beneficiary's SSN", "Benefits Paid", "Prior Year",
    "Benefits Repaid to SSA", "Net Benefits", "Voluntary Federal Withholding",
    "Total Lump Sum Social Security Received", "Lump Sum Taxable Social Security",
    "Medicare Premiums Withheld", "Prescription Drug Coverage Insurance",
    "Tier 1 Railroad Retirement", "Tier 1 Railroad Retirement Repaid",
    "Section 931 and 933 Exclusions", "FS", "State", "City", "Foreign Country",
    "Foreign Income Category", "Total Foreign Days", "Base Days If Not 240",
    "Foreign Days Before", "Foreign Days After", "Percent Foreign",
    "Social Security Nontaxable According to Treaty Provision", "Dual Status",
    "Address Line 1", "Address Line 2", "Address Line 3", "Claim Number",
    "Business Name", "Farm Name", "Passthrough Activity Name", "Do Not Pro Forma"
]

# Key column positions (1-indexed) for SSA-1099
SSA_MAP = {
    'tsj': 1,
    'name': 2,
    'beneficiary_ssn': 3,
    'benefits_paid': 4,       # Box 3
    'benefits_repaid': 6,     # Box 4
    'net_benefits': 7,        # Box 5
    'fed_withholding': 8,     # Box 6
}

# W-2: Wages and Salaries Worksheet (86 columns)
W2_COLUMNS = [
    "TS", "Employer Name", "Wages", "Prior Year", "Federal Tax Withheld",
    "Social Security Wages", "Social Security Tax Withheld", "Medicare Wages",
    "Medicare Tax Withheld", "Social Security Tips", "Allocated Tips",
    "Dependent Care Benefits", "Nonqualified Plans", "Retirement Plan",
    "Third-Party Sick Pay", "FS", "Federal Emp. ID Number", "Employer Address",
    "Employer City", "Emp. State", "Emp. ZIP", "Emp. Foreign Country Code",
    "Foreign Employer's ID Number", "Foreign Employer's State or Province",
    "W-2s Non-Standard", "Country Code", "Control Number", "Stat. Employee",
    "Schedule C Name", "Name", "Address", "City", "State", "ZIP",
    "Country Code", "SSN", "Third-Party Employer's EIN", "Third-Party Employer's Name",
    "State", "State ID", "State Wages", "State Tax", "Local Wages", "Local Tax",
    "Locality", "State", "State ID", "State Wages", "State Tax", "Local Wages",
    "Local Tax", "Locality", "Box 12 Code", "Box 12 Amount", "Box 12 Code",
    "Box 12 Amount", "Box 12 Code", "Box 12 Amount", "Box 12 Code", "Box 12 Amount",
    "Box 14 Description", "Box 14 Amount", "Box 14 Description", "Box 14 Amount",
    "Box 14 Description", "Box 14 Amount", "Box 14 Description", "Box 14 Amount",
    "Box 14 Description", "Box 14 Amount", "Box 14 Description", "Box 14 Amount",
    "W-2 Type", "W2c Original Wages", "W2c Original Fed WH", "W2c Original SS Wages",
    "W2c Original SS Tax", "W2c Original Medicare Wages", "W2c Original Medicare Tax",
    "W2c Original SS Tips", "W2c Original Allocated Tips", "W2c Original Dependent Care",
    "W2c Original Nonqual Plans", "Combat Pay"
]

# Key column positions (1-indexed) for W-2
W2_MAP = {
    'ts': 1,                        # T=Taxpayer, S=Spouse
    'employer_name': 2,
    'wages': 3,                     # Box 1
    'fed_withholding': 5,           # Box 2
    'ss_wages': 6,                  # Box 3
    'ss_tax': 7,                    # Box 4
    'medicare_wages': 8,            # Box 5
    'medicare_tax': 9,              # Box 6
    'employer_ein': 17,
    'state': 39,
    'state_id': 40,
    'state_wages': 41,              # Box 16
    'state_tax': 42,                # Box 17
}

# 1098: Mortgage Interest Worksheet (35 columns)
MTG_COLUMNS = [
    "TSJ", "FS", "ST", "Recipient's/Lender's Name", "Recipient's Address",
    "Recipient's City", "Recipient's State", "Recipient's ZIP Code",
    "Recipient's Foreign Country Code", "Recipient's Federal Identification Number",
    "Recipient's Phone Number", "Mortgage Interest", "Refund of Overpaid Interest",
    "Points Paid", "Taxes", "Payer's SSN", "Mortgage Interest Option",
    "Investment Interest", "Treat Premiums as Sch A Investment Expense (State Use)",
    "Payer's Name", "Payer's Address", "Payer's City", "Payer's State",
    "Payer's ZIP Code", "Outstanding Mortgage Principal", "Mortgage Origination Date",
    "Mortgage Insurance Premiums", "Account Number", "Number of Properties",
    "Property Address Same As Payer", "Address or Description of Property",
    "Additional  Address or Description of Property",
    "Additional  Address or Description of Property",
    "Additional  Address or Description of Property", "Mortgage Acquisition Date"
]

# Key column positions (1-indexed) for 1098
MTG_MAP = {
    'tsj': 1,
    'lender_name': 4,               # Recipient's/Lender's Name
    'lender_address': 5,
    'lender_city': 6,
    'lender_state': 7,
    'lender_zip': 8,
    'lender_ein': 10,               # Recipient's Federal ID
    'mortgage_interest': 12,        # Box 1
    'points_paid': 14,              # Box 6
    'property_tax': 15,             # Taxes (Box 10)
    'outstanding_principal': 25,    # Box 2
    'mortgage_insurance': 27,       # Box 5
    'property_address': 31,
}

# 1099-R: Distributions from Pensions, Annuities, IRAs (103 columns)
R_MAP = {
    'tsj': 1,
    'payer_name': 2,
    'gross_distribution': 3,        # Box 1
    'taxable_amount': 5,            # Box 2a
    'capital_gain': 6,              # Box 3
    'fed_withholding': 7,           # Box 4
    'state_withholding': 8,
    'local_withholding': 9,
    'dist_code': 10,                # Box 7
    'ira_sep_simple': 12,
    'payer_state_id': 16,
    'payer_ein': 19,
    'payer_address': 20,
    'payer_city': 21,
    'payer_state': 22,
    'payer_zip': 23,
}

# 1099-B: Capital Gains and Losses (53 columns)
B_MAP = {
    'description': 1,
    'quantity': 2,
    'sales_price': 3,               # Proceeds
    'cost_basis': 4,
    'date_acquired': 6,
    'date_sold': 7,
    'term_code': 8,                 # S=Short, L=Long
    'code_1099b': 9,                # A/B/D/E
    'fed_withholding': 20,
}

# 1099-NEC: Nonemployee Compensation (60 columns)
NEC_MAP = {
    'tsj': 1,
    'payer_name': 2,
    'payer_street': 3,
    'payer_city': 4,
    'payer_state': 5,
    'payer_zip': 6,
    'payer_ein': 11,
    'recipient_tin': 12,
    'recipient_name': 13,
    'recipient_street': 14,
    'recipient_city': 15,
    'recipient_state': 16,
    'recipient_zip': 17,
    'nec_compensation': 22,         # Box 1
    'fed_withholding': 26,          # Box 4
    'state_withholding': 27,
    'state_income': 29,
}

# 1099-G: Certain Government Payments (54 columns)
G_MAP = {
    'tsj': 1,
    'payer_name': 2,
    'payer_street': 3,
    'payer_city': 4,
    'payer_state': 5,
    'payer_zip': 6,
    'payer_ein': 8,
    'recipient_tin': 9,
    'recipient_name': 10,
    'unemployment': 18,             # Box 1
    'state_tax_refund': 20,         # Box 2
    'fed_withholding': 24,          # Box 4
    'rtaa_payments': 25,            # Box 5
    'taxable_grants': 26,           # Box 6
    'agriculture': 27,              # Box 7
}

# 1099-MISC: Miscellaneous Income (72 columns)
MISC_MAP = {
    'tsj': 1,
    'payer_name': 2,
    'payer_street': 3,
    'payer_city': 4,
    'payer_state': 5,
    'payer_zip': 6,
    'payer_tin': 10,
    'recipient_tin': 11,
    'recipient_name': 12,
    'rents': 24,                    # Box 1
    'royalties': 25,                # Box 2
    'other_income': 26,             # Box 3
    'fed_withholding': 27,          # Box 4
    'fishing_boat': 28,             # Box 5
    'medical_payments': 29,         # Box 6
    'substitute_payments': 31,      # Box 8
    'crop_insurance': 32,           # Box 9
    'gross_proceeds_attorney': 33,  # Box 10
    'fish_purchased': 34,           # Box 11
    'excess_golden_parachute': 36,  # Box 13
}

# 1099-K: Payment Card Transactions (33 columns)
K_MAP = {
    'tsj': 1,
    'filer_name': 2,
    'filer_address': 3,
    'filer_city': 4,
    'filer_state': 5,
    'filer_zip': 6,
    'filer_tin': 10,
    'payee_tin': 11,
    'payee_name': 12,
    'payee_address': 13,
    'gross_amount': 19,             # Box 1a
    'card_not_present': 20,         # Box 1b
    'fed_withholding': 23,
}

# 1098-E: Student Loan Interest (27 columns)
SLI_MAP = {
    'tsj': 1,
    'lender_name': 2,
    'lender_address': 3,
    'lender_city': 4,
    'lender_state': 5,
    'lender_zip': 6,
    'lender_ein': 10,
    'borrower_ssn': 11,
    'borrower_name': 12,
    'borrower_address': 13,
    'student_loan_interest': 20,    # Box 1
}

# 1099-SA: HSA/MSA Distributions (41 columns)
SA_MAP = {
    'ts': 1,
    'payer_name': 5,
    'payer_address': 6,
    'payer_city': 7,
    'payer_state': 8,
    'payer_zip': 9,
    'payer_ein': 13,
    'recipient_ssn': 14,
    'recipient_name': 15,
    'gross_distribution': 23,       # Box 1
    'earnings_excess': 24,          # Box 2
    'distribution_code': 26,        # Box 3
    'fmv_date_death': 25,
    'hsa_msa_archer': 27,           # Box 5
}

# 1098-T Tuition Statement (32 columns)
T_MAP = {
    'tsj': 1,
    'school_name': 2,
    'school_address': 3,
    'school_city': 4,
    'school_state': 5,
    'school_zip': 6,
    'school_ein': 10,
    'student_ssn': 11,
    'student_name': 12,
    'payments_received': 20,        # Box 1
    'amounts_billed': 21,           # Box 2 (older forms)
    'adjustments_prior': 22,        # Box 4
    'scholarships': 23,             # Box 5
    'adjustments_scholarships': 24, # Box 6
    'box7_checked': 25,
    'half_time': 26,                # Box 8
    'graduate': 27,                 # Box 9
}

# 1099-Q Qualified Education Programs (35 columns)
Q_MAP = {
    'tsj': 1,
    'payer_name': 2,
    'payer_address': 3,
    'payer_city': 4,
    'payer_state': 5,
    'payer_zip': 6,
    'payer_ein': 10,
    'recipient_ssn': 11,
    'recipient_name': 12,
    'gross_distribution': 20,       # Box 1
    'earnings': 21,                 # Box 2
    'basis': 22,                    # Box 3
    'trustee_transfer': 23,         # Box 4
    'distribution_type': 24,        # Box 5 (1=529, 2=Coverdell)
    'designated_beneficiary': 25,   # Box 6
}

# K-1 Partnership (99 columns)
K1_MAP = {
    'client_id': 1,
    'entity_name': 2,
    'entity_ein': 3,
    'partner_name': 4,
    'partner_ssn': 5,
    'activity_name': 8,
    'ordinary_income': 13,          # Line 1
    'net_rental_re': 14,            # Line 2
    'other_rental': 15,             # Line 3
    'guaranteed_services': 16,      # Line 4a
    'guaranteed_capital': 17,       # Line 4b
    'interest': 19,                 # Line 5
    'ordinary_dividends': 21,       # Line 6a
    'qualified_dividends': 22,      # Line 6b
    'royalties': 23,                # Line 7
    'st_capital_gl': 24,            # Line 8
    'lt_capital_gl': 25,            # Line 9a
    'section_1231_gl': 28,          # Line 10
}


def create_cch_workbook(sheet_name, columns, client_id=None):
    """Create a CCH-formatted workbook with headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Row 1: Client ID (optional - CCH adds this on export)
    if client_id:
        ws.cell(row=1, column=1, value=client_id)

    # Row 2: Worksheet path (optional)
    ws.cell(row=2, column=1, value=f"{sheet_name} - {sheet_name}")

    # Row 3: blank

    # Row 4: Section header
    ws.cell(row=4, column=1, value=sheet_name)

    # Row 5: blank (grouping row in CCH)

    # Row 6: Column headers
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.font = Font(bold=True)

    return wb, ws


# Minimum confidence threshold for including in CCH imports
MIN_CONFIDENCE_THRESHOLD = 60


def filter_quality_entries(data_list, form_type):
    """
    Filter out entries with quality issues (math errors, low confidence, garbage data).

    Args:
        data_list: List of parsed form entries
        form_type: Form type for logging

    Returns:
        List of entries that pass quality checks
    """
    if not data_list:
        return []

    filtered = []
    skipped = 0

    for item in data_list:
        quality = item.get('_quality', {})

        # Skip if overall confidence is too low
        overall = quality.get('overall_confidence', 100)
        if overall < MIN_CONFIDENCE_THRESHOLD:
            skipped += 1
            continue

        # Skip if there are math errors (indicates wrong values parsed)
        math_errors = quality.get('math_errors', [])
        if math_errors:
            skipped += 1
            continue

        # Skip if payer/employer name looks like garbage (form labels, etc.)
        name = item.get('employer_name', '') or item.get('payer_name', '') or ''
        name_lower = name.lower()
        garbage_patterns = ['zip', '1099', 'box ', 'federal', 'withheld', 'income tax',
                            'postal code', 'telephone', 'payer', 'recipient', 'form w-2',
                            'fed.', 'medicare', 'social security', 'wages', 'w-2 box']
        is_garbage = False
        for pattern in garbage_patterns:
            if pattern in name_lower:
                is_garbage = True
                break
        if is_garbage:
            skipped += 1
            continue

        # Skip if interest amount looks like a tax year
        interest = item.get('box1_interest', 0)
        if interest in [2023, 2024, 2025, 2026]:
            skipped += 1
            continue

        filtered.append(item)

    if skipped:
        print(f"    ({skipped} {form_type} entries skipped due to quality issues)")

    return filtered


def generate_1099int_import(data_list, output_path, client_id=None):
    """
    Generate CCH-compatible Excel for 1099-INT import.

    Args:
        data_list: List of dicts from parse_source_docs.py with keys:
            - payer_name: str
            - box1_interest: float
            - box3_savings_bond: float (optional)
            - box4_fed_withholding: float
            - box8_tax_exempt_interest: float (optional)
            - payer_tin: str (optional, format: XX-XXXXXXX)
        output_path: Where to save the Excel file
        client_id: Optional CCH client ID string
    """
    wb, ws = create_cch_workbook("Interest", INT_COLUMNS, client_id)

    # Write data rows starting at row 7
    for row_idx, entry in enumerate(data_list, 7):
        # TSJ - leave blank for joint, or set T/S
        ws.cell(row=row_idx, column=INT_MAP['tsj'], value="")

        # Payer name
        ws.cell(row=row_idx, column=INT_MAP['payer'],
                value=entry.get('payer_name', ''))

        # Box 1: Interest Income
        interest = entry.get('box1_interest', 0)
        if interest:
            ws.cell(row=row_idx, column=INT_MAP['interest_income'], value=interest)

        # Box 3: U.S. Savings Bonds
        savings = entry.get('box3_savings_bond', 0)
        if savings:
            ws.cell(row=row_idx, column=INT_MAP['us_savings_bonds'], value=savings)

        # Box 4: Federal Tax Withheld
        fed_wh = entry.get('box4_fed_withholding', 0)
        if fed_wh:
            ws.cell(row=row_idx, column=INT_MAP['fed_withholding'], value=fed_wh)

        # Box 8: Tax-Exempt Interest
        tax_exempt = entry.get('box8_tax_exempt_interest', 0)
        if tax_exempt:
            ws.cell(row=row_idx, column=INT_MAP['tax_exempt_interest'], value=tax_exempt)

        # EIN
        ein = entry.get('payer_tin', '')
        if ein:
            ws.cell(row=row_idx, column=INT_MAP['ein'], value=ein)

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} 1099-INT entries")
    return output_path


def generate_1099div_import(data_list, output_path, client_id=None):
    """
    Generate CCH-compatible Excel for 1099-DIV import.

    Args:
        data_list: List of dicts from parse_source_docs.py with keys:
            - payer_name: str
            - box1a_ordinary_dividends: float
            - box1b_qualified_dividends: float
            - box2a_total_cap_gain: float
            - box3_nondiv_dist: float
            - box4_fed_withholding: float
            - box5_sec199a: float
            - box7_foreign_tax: float
            - box12_exempt_int_div: float
            - payer_tin: str (optional)
    """
    wb, ws = create_cch_workbook("Dividends", DIV_COLUMNS, client_id)

    for row_idx, entry in enumerate(data_list, 7):
        ws.cell(row=row_idx, column=DIV_MAP['tsj'], value="")

        ws.cell(row=row_idx, column=DIV_MAP['payer_name'],
                value=entry.get('payer_name', ''))

        # Box 1a: Ordinary Dividends
        ord_div = entry.get('box1a_ordinary_dividends', 0)
        if ord_div:
            ws.cell(row=row_idx, column=DIV_MAP['ordinary_dividends'], value=ord_div)

        # Box 1b: Qualified Dividends
        qual_div = entry.get('box1b_qualified_dividends', 0)
        if qual_div:
            ws.cell(row=row_idx, column=DIV_MAP['qualified_dividends'], value=qual_div)

        # Box 2a: Total Capital Gain Distribution
        cap_gain = entry.get('box2a_total_cap_gain', 0)
        if cap_gain:
            ws.cell(row=row_idx, column=DIV_MAP['total_cap_gain_dist'], value=cap_gain)

        # Box 3: Nondividend Distributions
        nondiv = entry.get('box3_nondiv_dist', 0)
        if nondiv:
            ws.cell(row=row_idx, column=DIV_MAP['nondividend_dist'], value=nondiv)

        # Box 4: Federal Tax Withheld
        fed_wh = entry.get('box4_fed_withholding', 0)
        if fed_wh:
            ws.cell(row=row_idx, column=DIV_MAP['fed_withholding'], value=fed_wh)

        # Box 5: Section 199A Dividends (column 75!)
        sec199a = entry.get('box5_sec199a', 0)
        if sec199a:
            ws.cell(row=row_idx, column=DIV_MAP['section_199a_dividends'], value=sec199a)

        # Box 7: Foreign Tax Paid
        foreign = entry.get('box7_foreign_tax', 0)
        if foreign:
            ws.cell(row=row_idx, column=DIV_MAP['foreign_tax_paid'], value=foreign)

        # Box 12: Exempt-Interest Dividends
        exempt = entry.get('box12_exempt_int_div', 0)
        if exempt:
            ws.cell(row=row_idx, column=DIV_MAP['exempt_int_dividends'], value=exempt)

        # EIN
        ein = entry.get('payer_tin', '')
        if ein:
            ws.cell(row=row_idx, column=DIV_MAP['ein'], value=ein)

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} 1099-DIV entries")
    return output_path


def generate_ssa1099_import(data_list, output_path, client_id=None):
    """
    Generate CCH-compatible Excel for SSA-1099 import.

    Args:
        data_list: List of dicts from parse_source_docs.py with keys:
            - description: str (beneficiary name)
            - beneficiary_ssn: str (optional)
            - box3_benefits_paid: float
            - box4_benefits_repaid: float (optional)
            - box5_net_benefits: float
            - box6_fed_withholding: float
    """
    wb, ws = create_cch_workbook("Social Security Benefit Stmt", SSA_COLUMNS, client_id)

    for row_idx, entry in enumerate(data_list, 7):
        ws.cell(row=row_idx, column=SSA_MAP['tsj'], value="")

        # Name (from description or dedicated field)
        # Use placeholder if name is generic or missing
        name = entry.get('beneficiary_name', entry.get('description', ''))
        if not name or name == 'Social Security Benefits':
            beneficiary_num = row_idx - 6  # 1-indexed
            name = f"Beneficiary {beneficiary_num}"
        ws.cell(row=row_idx, column=SSA_MAP['name'], value=name)

        # Beneficiary SSN
        ssn = entry.get('beneficiary_ssn', '')
        if ssn:
            ws.cell(row=row_idx, column=SSA_MAP['beneficiary_ssn'], value=ssn)

        # Box 3: Benefits Paid
        benefits_paid = entry.get('box3_benefits_paid', 0)
        if benefits_paid:
            ws.cell(row=row_idx, column=SSA_MAP['benefits_paid'], value=benefits_paid)

        # Box 4: Benefits Repaid
        repaid = entry.get('box4_benefits_repaid', 0)
        if repaid:
            ws.cell(row=row_idx, column=SSA_MAP['benefits_repaid'], value=repaid)

        # Box 5: Net Benefits
        net = entry.get('box5_net_benefits', 0)
        if net:
            ws.cell(row=row_idx, column=SSA_MAP['net_benefits'], value=net)

        # Box 6: Federal Withholding
        fed_wh = entry.get('box6_fed_withholding', 0)
        if fed_wh:
            ws.cell(row=row_idx, column=SSA_MAP['fed_withholding'], value=fed_wh)

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} SSA-1099 entries")
    return output_path


def generate_w2_import(data_list, output_path, client_id=None):
    """
    Generate CCH-compatible Excel for W-2 import.

    Args:
        data_list: List of dicts from parse_source_docs.py with keys:
            - employer_name: str
            - box1_wages: float
            - box2_fed_withholding: float
            - box3_ss_wages: float
            - box4_ss_tax: float
            - box5_medicare_wages: float
            - box6_medicare_tax: float
            - box16_state_wages: float
            - box17_state_withholding: float
            - employer_ein: str (optional)
    """
    wb, ws = create_cch_workbook("Wages and Salaries", W2_COLUMNS, client_id)

    for row_idx, entry in enumerate(data_list, 7):
        # TS - T=Taxpayer, S=Spouse (leave blank or set based on filename)
        ts = entry.get('ts', '')
        ws.cell(row=row_idx, column=W2_MAP['ts'], value=ts)

        # Employer Name
        ws.cell(row=row_idx, column=W2_MAP['employer_name'],
                value=entry.get('employer_name', ''))

        # Box 1: Wages
        wages = entry.get('box1_wages', 0)
        if wages:
            ws.cell(row=row_idx, column=W2_MAP['wages'], value=wages)

        # Box 2: Federal Tax Withheld
        fed_wh = entry.get('box2_fed_withholding', 0)
        if fed_wh:
            ws.cell(row=row_idx, column=W2_MAP['fed_withholding'], value=fed_wh)

        # Box 3: Social Security Wages
        ss_wages = entry.get('box3_ss_wages', 0)
        if ss_wages:
            ws.cell(row=row_idx, column=W2_MAP['ss_wages'], value=ss_wages)

        # Box 4: Social Security Tax
        ss_tax = entry.get('box4_ss_tax', 0)
        if ss_tax:
            ws.cell(row=row_idx, column=W2_MAP['ss_tax'], value=ss_tax)

        # Box 5: Medicare Wages
        med_wages = entry.get('box5_medicare_wages', 0)
        if med_wages:
            ws.cell(row=row_idx, column=W2_MAP['medicare_wages'], value=med_wages)

        # Box 6: Medicare Tax
        med_tax = entry.get('box6_medicare_tax', 0)
        if med_tax:
            ws.cell(row=row_idx, column=W2_MAP['medicare_tax'], value=med_tax)

        # Employer EIN
        ein = entry.get('employer_ein', '')
        if ein:
            ws.cell(row=row_idx, column=W2_MAP['employer_ein'], value=ein)

        # Box 16: State Wages
        state_wages = entry.get('box16_state_wages', 0)
        if state_wages:
            ws.cell(row=row_idx, column=W2_MAP['state_wages'], value=state_wages)

        # Box 17: State Tax Withheld
        state_tax = entry.get('box17_state_withholding', 0)
        if state_tax:
            ws.cell(row=row_idx, column=W2_MAP['state_tax'], value=state_tax)

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} W-2 entries")
    return output_path


def generate_1098_import(data_list, output_path, client_id=None):
    """
    Generate CCH-compatible Excel for 1098 Mortgage Interest import.

    Args:
        data_list: List of dicts from parse_source_docs.py with keys:
            - lender_name: str
            - box1_mortgage_interest: float
            - box2_outstanding_principal: float
            - box5_mortgage_insurance: float
            - box10_property_tax: float
            - property_address: str
            - lender_ein: str (optional)
    """
    wb, ws = create_cch_workbook("Mortgage Interest", MTG_COLUMNS, client_id)

    for row_idx, entry in enumerate(data_list, 7):
        ws.cell(row=row_idx, column=MTG_MAP['tsj'], value="")

        # Lender Name
        ws.cell(row=row_idx, column=MTG_MAP['lender_name'],
                value=entry.get('lender_name', ''))

        # Lender EIN
        ein = entry.get('lender_ein', '')
        if ein:
            ws.cell(row=row_idx, column=MTG_MAP['lender_ein'], value=ein)

        # Box 1: Mortgage Interest
        interest = entry.get('box1_mortgage_interest', 0)
        if interest:
            ws.cell(row=row_idx, column=MTG_MAP['mortgage_interest'], value=interest)

        # Box 2: Outstanding Principal
        principal = entry.get('box2_outstanding_principal', 0)
        if principal:
            ws.cell(row=row_idx, column=MTG_MAP['outstanding_principal'], value=principal)

        # Box 5: Mortgage Insurance Premiums
        insurance = entry.get('box5_mortgage_insurance', 0)
        if insurance:
            ws.cell(row=row_idx, column=MTG_MAP['mortgage_insurance'], value=insurance)

        # Box 10: Property Tax
        prop_tax = entry.get('box10_property_tax', 0)
        if prop_tax:
            ws.cell(row=row_idx, column=MTG_MAP['property_tax'], value=prop_tax)

        # Property Address
        address = entry.get('property_address', '')
        if address:
            ws.cell(row=row_idx, column=MTG_MAP['property_address'], value=address)

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} 1098 entries")
    return output_path


def generate_1099r_import(data_list, output_path, client_id=None):
    """Generate CCH-compatible Excel for 1099-R import."""
    wb, ws = create_cch_workbook("Dist Pensions Annuities IRAs", [], client_id)

    # Write headers (simplified - just the key ones we use)
    headers = ["TSJ", "Payer's Name", "Gross Distribution", "Prior Year", "Taxable Amount",
               "Capital Gain", "Federal Tax Withheld", "State Tax Withheld", "Local Tax Withheld",
               "Dist. Code", "IRA", "IRA/SEP/SIMPLE", "FS", "State", "City",
               "Payer's State ID Number", "Date of Payment", "Acct. Number", "Payer's Federal ID Number"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=6, column=col_idx, value=header)

    for row_idx, entry in enumerate(data_list, 7):
        ws.cell(row=row_idx, column=R_MAP['tsj'], value=entry.get('tsj', ''))
        ws.cell(row=row_idx, column=R_MAP['payer_name'], value=entry.get('payer_name', ''))

        if entry.get('box1_gross_distribution'):
            ws.cell(row=row_idx, column=R_MAP['gross_distribution'], value=entry['box1_gross_distribution'])
        if entry.get('box2a_taxable_amount'):
            ws.cell(row=row_idx, column=R_MAP['taxable_amount'], value=entry['box2a_taxable_amount'])
        if entry.get('box4_fed_withholding'):
            ws.cell(row=row_idx, column=R_MAP['fed_withholding'], value=entry['box4_fed_withholding'])
        if entry.get('box7_distribution_code'):
            ws.cell(row=row_idx, column=R_MAP['dist_code'], value=entry['box7_distribution_code'])
        if entry.get('payer_ein'):
            ws.cell(row=row_idx, column=R_MAP['payer_ein'], value=entry['payer_ein'])

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} 1099-R entries")
    return output_path


def generate_1099b_import(data_list, output_path, client_id=None):
    """Generate CCH-compatible Excel for 1099-B Capital Gains import."""
    wb, ws = create_cch_workbook("Capital Gains and Losses", [], client_id)

    headers = ["Description", "Quantity", "Sales Price", "Cost or Other Basis",
               "Accountant Gain / Loss - Override", "Date Acquired", "Date Sold",
               "Term Code", "1099-B Code", "Corrected 1099-B Basis"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=6, column=col_idx, value=header)

    for row_idx, entry in enumerate(data_list, 7):
        ws.cell(row=row_idx, column=B_MAP['description'], value=entry.get('description', ''))
        if entry.get('quantity'):
            ws.cell(row=row_idx, column=B_MAP['quantity'], value=entry['quantity'])
        if entry.get('proceeds') or entry.get('sales_price'):
            ws.cell(row=row_idx, column=B_MAP['sales_price'], value=entry.get('proceeds', entry.get('sales_price')))
        if entry.get('cost_basis'):
            ws.cell(row=row_idx, column=B_MAP['cost_basis'], value=entry['cost_basis'])
        if entry.get('date_acquired'):
            ws.cell(row=row_idx, column=B_MAP['date_acquired'], value=entry['date_acquired'])
        if entry.get('date_sold'):
            ws.cell(row=row_idx, column=B_MAP['date_sold'], value=entry['date_sold'])
        if entry.get('term_code'):
            ws.cell(row=row_idx, column=B_MAP['term_code'], value=entry['term_code'])
        if entry.get('code_1099b'):
            ws.cell(row=row_idx, column=B_MAP['code_1099b'], value=entry['code_1099b'])

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} 1099-B entries")
    return output_path


def generate_1099nec_import(data_list, output_path, client_id=None):
    """Generate CCH-compatible Excel for 1099-NEC import."""
    wb, ws = create_cch_workbook("Nonemployee Compensation", [], client_id)

    headers = ["TSJ", "Name", "Street", "City", "State", "ZIP Code", "Foreign Country",
               "Province, State or County", "Postal Code", "Telephone Number",
               "Federal Identification Number", "Identification Number", "Name", "Street",
               "City", "State", "ZIP or Postal Code", "Foreign Country", "Province, State or County",
               "Account Number", "2nd TIN Not.", "Nonemployee Compensation"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=6, column=col_idx, value=header)

    for row_idx, entry in enumerate(data_list, 7):
        ws.cell(row=row_idx, column=NEC_MAP['tsj'], value=entry.get('tsj', ''))
        ws.cell(row=row_idx, column=NEC_MAP['payer_name'], value=entry.get('payer_name', ''))
        if entry.get('payer_ein'):
            ws.cell(row=row_idx, column=NEC_MAP['payer_ein'], value=entry['payer_ein'])
        if entry.get('box1_nec') or entry.get('nonemployee_compensation'):
            ws.cell(row=row_idx, column=NEC_MAP['nec_compensation'],
                    value=entry.get('box1_nec', entry.get('nonemployee_compensation')))
        if entry.get('box4_fed_withholding'):
            ws.cell(row=row_idx, column=NEC_MAP['fed_withholding'], value=entry['box4_fed_withholding'])

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} 1099-NEC entries")
    return output_path


def generate_1099g_import(data_list, output_path, client_id=None):
    """Generate CCH-compatible Excel for 1099-G import."""
    wb, ws = create_cch_workbook("Certain Government Payments", [], client_id)

    headers = ["TSJ", "Name", "Street", "City", "State", "ZIP Code", "Telephone Number",
               "Federal Identification Number", "Identification Number", "Name", "Address",
               "City", "State", "ZIP or Postal Code", "Foreign Country", "Province / State / County",
               "Account Number", "Unemployment Compensation", "Compensation Repaid",
               "Box 2 State Income Tax Refund", "Box 2 Local Income Tax Refund", "Prior Year",
               "Year for Box 2", "Federal Income Tax Withheld"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=6, column=col_idx, value=header)

    for row_idx, entry in enumerate(data_list, 7):
        ws.cell(row=row_idx, column=G_MAP['tsj'], value=entry.get('tsj', ''))
        ws.cell(row=row_idx, column=G_MAP['payer_name'], value=entry.get('payer_name', ''))
        if entry.get('box1_unemployment'):
            ws.cell(row=row_idx, column=G_MAP['unemployment'], value=entry['box1_unemployment'])
        if entry.get('box2_state_refund'):
            ws.cell(row=row_idx, column=G_MAP['state_tax_refund'], value=entry['box2_state_refund'])
        if entry.get('box4_fed_withholding'):
            ws.cell(row=row_idx, column=G_MAP['fed_withholding'], value=entry['box4_fed_withholding'])

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} 1099-G entries")
    return output_path


def generate_1099misc_import(data_list, output_path, client_id=None):
    """Generate CCH-compatible Excel for 1099-MISC import."""
    wb, ws = create_cch_workbook("Miscellaneous Information", [], client_id)

    headers = ["TSJ", "Name", "Street", "City", "State", "ZIP or Postal Code",
               "Foreign Country", "Province, State or County", "Telephone Number",
               "Payer's TIN", "Recipient's TIN", "Name", "Street", "City", "State",
               "ZIP or Postal Code", "Foreign Country", "Province, State or County",
               "Account Number", "FATCA Filing Requirement", "2nd TIN Not.",
               "Section 409A Deferrals", "Section 409A Income", "Rents", "Royalties",
               "Other Income", "Federal Tax Withheld"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=6, column=col_idx, value=header)

    for row_idx, entry in enumerate(data_list, 7):
        ws.cell(row=row_idx, column=MISC_MAP['tsj'], value=entry.get('tsj', ''))
        ws.cell(row=row_idx, column=MISC_MAP['payer_name'], value=entry.get('payer_name', ''))
        if entry.get('box1_rents'):
            ws.cell(row=row_idx, column=MISC_MAP['rents'], value=entry['box1_rents'])
        if entry.get('box2_royalties'):
            ws.cell(row=row_idx, column=MISC_MAP['royalties'], value=entry['box2_royalties'])
        if entry.get('box3_other_income'):
            ws.cell(row=row_idx, column=MISC_MAP['other_income'], value=entry['box3_other_income'])
        if entry.get('box4_fed_withholding'):
            ws.cell(row=row_idx, column=MISC_MAP['fed_withholding'], value=entry['box4_fed_withholding'])

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} 1099-MISC entries")
    return output_path


def generate_1099k_import(data_list, output_path, client_id=None):
    """Generate CCH-compatible Excel for 1099-K import."""
    wb, ws = create_cch_workbook("Payment Card Transactions", [], client_id)

    headers = ["TSJ", "Name", "Address", "City", "State", "ZIP or Postal Code",
               "Foreign Country", "Province / County", "Telephone Number", "TIN", "TIN",
               "Name", "Address", "City", "State", "ZIP or Postal Code", "Foreign Country",
               "Province / County", "Gross Amount of Payment Card / Third Party Network Transactions",
               "Card Not Present Transactions", "Cash Tips", "Treasury Tipped Occupation Code",
               "Federal Income Tax Withheld"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=6, column=col_idx, value=header)

    for row_idx, entry in enumerate(data_list, 7):
        ws.cell(row=row_idx, column=K_MAP['tsj'], value=entry.get('tsj', ''))
        ws.cell(row=row_idx, column=K_MAP['filer_name'], value=entry.get('filer_name', entry.get('payer_name', '')))
        if entry.get('payee_name'):
            ws.cell(row=row_idx, column=K_MAP['payee_name'], value=entry['payee_name'])
        if entry.get('box1a_gross_amount'):
            ws.cell(row=row_idx, column=K_MAP['gross_amount'], value=entry['box1a_gross_amount'])
        if entry.get('fed_withholding'):
            ws.cell(row=row_idx, column=K_MAP['fed_withholding'], value=entry['fed_withholding'])

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} 1099-K entries")
    return output_path


def generate_1098e_import(data_list, output_path, client_id=None):
    """Generate CCH-compatible Excel for 1098-E Student Loan Interest import."""
    wb, ws = create_cch_workbook("Student Loan Interest Statement", [], client_id)

    headers = ["TSJ", "Name", "Address", "City", "State", "ZIP or Postal Code",
               "Province / State / County", "Foreign Country", "Telephone Number",
               "Federal Identification Number", "Social Security Number", "Name",
               "Address", "City", "State", "ZIP or Postal Code", "Province / State / County",
               "Foreign Country", "Account Number", "Student Loan Interest Received"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=6, column=col_idx, value=header)

    for row_idx, entry in enumerate(data_list, 7):
        ws.cell(row=row_idx, column=SLI_MAP['tsj'], value=entry.get('tsj', ''))
        ws.cell(row=row_idx, column=SLI_MAP['lender_name'], value=entry.get('lender_name', ''))
        if entry.get('lender_ein'):
            ws.cell(row=row_idx, column=SLI_MAP['lender_ein'], value=entry['lender_ein'])
        if entry.get('borrower_name'):
            ws.cell(row=row_idx, column=SLI_MAP['borrower_name'], value=entry['borrower_name'])
        if entry.get('box1_interest'):
            ws.cell(row=row_idx, column=SLI_MAP['student_loan_interest'], value=entry['box1_interest'])

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} 1098-E entries")
    return output_path


def generate_1099sa_import(data_list, output_path, client_id=None):
    """Generate CCH-compatible Excel for 1099-SA HSA/MSA import."""
    wb, ws = create_cch_workbook("Distributions From HSA or MSA", [], client_id)

    headers = ["TS", "FS", "State", "City", "Payer Name", "Payer Address", "Payer City",
               "Payer State", "Payer ZIP code", "Payer Province", "Payer Foreign Country",
               "Payer Phone Number", "Payer Federal ID", "Recipient ID", "Recipient Name",
               "Recipient Address", "Recipient City", "Recipient State", "Recipient ZIP Code",
               "Recipient Province", "Recipient Foreign Country", "Account Number",
               "Gross Distribution", "Earnings on Excess", "FMV on Date of Death",
               "Distribution Code", "HSA / Archer MSA / MA MSA"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=6, column=col_idx, value=header)

    for row_idx, entry in enumerate(data_list, 7):
        ws.cell(row=row_idx, column=SA_MAP['ts'], value=entry.get('ts', ''))
        ws.cell(row=row_idx, column=SA_MAP['payer_name'], value=entry.get('payer_name', ''))
        if entry.get('payer_ein'):
            ws.cell(row=row_idx, column=SA_MAP['payer_ein'], value=entry['payer_ein'])
        if entry.get('recipient_name'):
            ws.cell(row=row_idx, column=SA_MAP['recipient_name'], value=entry['recipient_name'])
        if entry.get('box1_gross_distribution'):
            ws.cell(row=row_idx, column=SA_MAP['gross_distribution'], value=entry['box1_gross_distribution'])
        if entry.get('box3_distribution_code'):
            ws.cell(row=row_idx, column=SA_MAP['distribution_code'], value=entry['box3_distribution_code'])

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} 1099-SA entries")
    return output_path


def generate_k1_import(data_list, output_path, client_id=None):
    """Generate CCH-compatible Excel for K-1 Partnership import."""
    wb, ws = create_cch_workbook("K-1 Activities (Federal)", [], client_id)

    headers = ["Import Client ID", "Passthrough Entity Name", "Passthrough Entity ID",
               "Partner / Shareholder / Beneficiary Name", "Partner / Shareholder / Beneficiary ID",
               "Preparer Notes", "Activity number", "Activity Name / Description",
               "100% Disposition ", "PTP (N/A 1041)", "Class code: 1 NP, 2 Act rental, 3 Passive, 4 MPREA",
               "Type of Property", "L1 Ordinary Income (Loss)", "L2 Net Rental Real Estate",
               "L3 Net Other Rent", "L4a Guaranteed Payments for Services",
               "L4b Guaranteed Payments for Capital", "L4c Total Guaranteed Payments",
               "L5 Interest", "Interest from US Bonds", "L6a Ordinary Dividends",
               "L6b Qualified Dividends", "L7 Royalties", "L8 Short-Term Capital G/L",
               "L9a Net Long-Term Capital G/L", "L9b Collectibles 28% G/L",
               "L9c Unrecaptured Sec 1250 gain", "L10 Section 1231 Gain (Loss)"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=6, column=col_idx, value=header)

    for row_idx, entry in enumerate(data_list, 7):
        ws.cell(row=row_idx, column=K1_MAP['entity_name'], value=entry.get('entity_name', ''))
        ws.cell(row=row_idx, column=K1_MAP['entity_ein'], value=entry.get('entity_ein', ''))
        ws.cell(row=row_idx, column=K1_MAP['partner_name'], value=entry.get('partner_name', ''))
        if entry.get('partner_ssn'):
            ws.cell(row=row_idx, column=K1_MAP['partner_ssn'], value=entry['partner_ssn'])
        if entry.get('box1_ordinary_income'):
            ws.cell(row=row_idx, column=K1_MAP['ordinary_income'], value=entry['box1_ordinary_income'])
        if entry.get('box2_net_rental_re'):
            ws.cell(row=row_idx, column=K1_MAP['net_rental_re'], value=entry['box2_net_rental_re'])
        if entry.get('box5_interest'):
            ws.cell(row=row_idx, column=K1_MAP['interest'], value=entry['box5_interest'])
        if entry.get('box6a_ordinary_dividends'):
            ws.cell(row=row_idx, column=K1_MAP['ordinary_dividends'], value=entry['box6a_ordinary_dividends'])
        if entry.get('box6b_qualified_dividends'):
            ws.cell(row=row_idx, column=K1_MAP['qualified_dividends'], value=entry['box6b_qualified_dividends'])
        if entry.get('box8_st_capital_gl'):
            ws.cell(row=row_idx, column=K1_MAP['st_capital_gl'], value=entry['box8_st_capital_gl'])
        if entry.get('box9a_lt_capital_gl'):
            ws.cell(row=row_idx, column=K1_MAP['lt_capital_gl'], value=entry['box9a_lt_capital_gl'])
        if entry.get('box10_section_1231'):
            ws.cell(row=row_idx, column=K1_MAP['section_1231_gl'], value=entry['box10_section_1231'])

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} K-1 entries")
    return output_path


def generate_1098t_import(data_list, output_path, client_id=None):
    """Generate CCH-compatible Excel for 1098-T Tuition Statement import."""
    wb, ws = create_cch_workbook("Tuition Statement", [], client_id)

    headers = ["TSJ", "Filer's Name", "Filer's Address", "Filer's City", "Filer's State",
               "Filer's ZIP Code", "Province / State / County", "Foreign Country",
               "Filer's Telephone Number", "Filer's Federal Identification Number",
               "Student's SSN", "Student's Name", "Student's Address", "Student's City",
               "Student's State", "Student's ZIP Code", "Student's Province",
               "Student's Foreign Country", "Account Number",
               "Payments Received for Qualified Tuition",  # Box 1
               "Amounts Billed for Qualified Tuition",      # Box 2
               "Adjustments Made for Prior Year",           # Box 4
               "Scholarships or Grants",                    # Box 5
               "Adjustments to Scholarships or Grants",     # Box 6
               "Checked if Box 1 Includes Jan-Mar",         # Box 7
               "At Least Half-Time Student",                # Box 8
               "Graduate Student"]                          # Box 9
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=6, column=col_idx, value=header)

    for row_idx, entry in enumerate(data_list, 7):
        ws.cell(row=row_idx, column=T_MAP['tsj'], value=entry.get('tsj', ''))
        ws.cell(row=row_idx, column=T_MAP['school_name'], value=entry.get('school_name', ''))

        if entry.get('school_ein'):
            ws.cell(row=row_idx, column=T_MAP['school_ein'], value=entry['school_ein'])
        if entry.get('student_name'):
            ws.cell(row=row_idx, column=T_MAP['student_name'], value=entry['student_name'])
        if entry.get('student_ssn'):
            ws.cell(row=row_idx, column=T_MAP['student_ssn'], value=entry['student_ssn'])

        # Box 1: Payments received
        if entry.get('box1_payments_received'):
            ws.cell(row=row_idx, column=T_MAP['payments_received'], value=entry['box1_payments_received'])

        # Box 2: Amounts billed (older forms)
        if entry.get('box2_amounts_billed'):
            ws.cell(row=row_idx, column=T_MAP['amounts_billed'], value=entry['box2_amounts_billed'])

        # Box 4: Adjustments for prior year
        if entry.get('box4_adjustments_prior_year'):
            ws.cell(row=row_idx, column=T_MAP['adjustments_prior'], value=entry['box4_adjustments_prior_year'])

        # Box 5: Scholarships or grants
        if entry.get('box5_scholarships'):
            ws.cell(row=row_idx, column=T_MAP['scholarships'], value=entry['box5_scholarships'])

        # Box 6: Adjustments to scholarships
        if entry.get('box6_adjustments_scholarships'):
            ws.cell(row=row_idx, column=T_MAP['adjustments_scholarships'], value=entry['box6_adjustments_scholarships'])

        # Box 7: Checked if amounts include Jan-Mar of next year
        if entry.get('box7_checked'):
            ws.cell(row=row_idx, column=T_MAP['box7_checked'], value='X')

        # Box 8: Half-time student
        if entry.get('box8_half_time'):
            ws.cell(row=row_idx, column=T_MAP['half_time'], value='X')

        # Box 9: Graduate student
        if entry.get('box9_graduate'):
            ws.cell(row=row_idx, column=T_MAP['graduate'], value='X')

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} 1098-T entries")
    return output_path


def generate_1099q_import(data_list, output_path, client_id=None):
    """Generate CCH-compatible Excel for 1099-Q Qualified Education import."""
    wb, ws = create_cch_workbook("Distributions from QEPs", [], client_id)

    headers = ["TSJ", "Payer/Trustee Name", "Payer Address", "Payer City", "Payer State",
               "Payer ZIP Code", "Province / State / County", "Foreign Country",
               "Telephone Number", "Federal Identification Number",
               "Recipient's TIN", "Recipient's Name", "Recipient's Address",
               "Recipient's City", "Recipient's State", "Recipient's ZIP Code",
               "Recipient's Province", "Recipient's Foreign Country", "Account Number",
               "Gross Distribution",                    # Box 1
               "Earnings",                              # Box 2
               "Basis",                                 # Box 3
               "Trustee-to-Trustee Transfer",          # Box 4
               "Distribution Type (1=529, 2=Coverdell)", # Box 5
               "Designated Beneficiary"]                # Box 6
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=6, column=col_idx, value=header)

    for row_idx, entry in enumerate(data_list, 7):
        ws.cell(row=row_idx, column=Q_MAP['tsj'], value=entry.get('tsj', ''))
        ws.cell(row=row_idx, column=Q_MAP['payer_name'], value=entry.get('payer_name', ''))

        if entry.get('payer_ein'):
            ws.cell(row=row_idx, column=Q_MAP['payer_ein'], value=entry['payer_ein'])
        if entry.get('recipient_name'):
            ws.cell(row=row_idx, column=Q_MAP['recipient_name'], value=entry['recipient_name'])
        if entry.get('recipient_ssn'):
            ws.cell(row=row_idx, column=Q_MAP['recipient_ssn'], value=entry['recipient_ssn'])

        # Box 1: Gross distribution
        if entry.get('box1_gross_distribution'):
            ws.cell(row=row_idx, column=Q_MAP['gross_distribution'], value=entry['box1_gross_distribution'])

        # Box 2: Earnings
        if entry.get('box2_earnings'):
            ws.cell(row=row_idx, column=Q_MAP['earnings'], value=entry['box2_earnings'])

        # Box 3: Basis
        if entry.get('box3_basis'):
            ws.cell(row=row_idx, column=Q_MAP['basis'], value=entry['box3_basis'])

        # Box 4: Trustee-to-trustee transfer
        if entry.get('box4_trustee_transfer'):
            ws.cell(row=row_idx, column=Q_MAP['trustee_transfer'], value='X')

        # Box 5: Distribution type (1=529, 2=Coverdell)
        if entry.get('box5_distribution_type'):
            ws.cell(row=row_idx, column=Q_MAP['distribution_type'], value=entry['box5_distribution_type'])

        # Box 6: Designated beneficiary
        if entry.get('box6_designated_beneficiary'):
            ws.cell(row=row_idx, column=Q_MAP['designated_beneficiary'], value='X')

    wb.save(output_path)
    print(f"  Created: {output_path}")
    print(f"    {len(data_list)} 1099-Q entries")
    return output_path


def has_ocr_entries(data_list):
    """Check if any entries in the list came from OCR."""
    for entry in data_list:
        quality = entry.get('_quality', {})
        if quality.get('is_ocr', False):
            return True
    return False


def has_validation_issues(data_list):
    """Check if any entries have validation issues requiring review."""
    for entry in data_list:
        quality = entry.get('_quality', {})
        # Check for issues
        if quality.get('issues'):
            return True
        # Check for missing required fields
        if quality.get('missing_required'):
            return True
        # Check for math errors
        if quality.get('math_errors'):
            return True
        # Check for low confidence (below 60%)
        if quality.get('overall_confidence', 100) < 60:
            return True
    return False


def get_import_filename(base_name, data_list):
    """
    Generate filename with appropriate suffix based on data quality.

    Returns filename like:
        CCH_Import_W2.xlsx - Clean digital
        CCH_Import_W2_OCR.xlsx - Has OCR entries
        CCH_Import_W2_REVIEW.xlsx - Has validation issues
        CCH_Import_W2_OCR_REVIEW.xlsx - Both
    """
    suffix = ""
    if has_ocr_entries(data_list):
        suffix += "_OCR"
    if has_validation_issues(data_list):
        suffix += "_REVIEW"
    return f"CCH_Import_{base_name}{suffix}.xlsx"


def generate_cch_imports(parsed_data, output_dir, forms=None, client_id=None):
    """
    Generate all CCH import files from parsed data.

    Args:
        parsed_data: Dict from parse_source_docs.py JSON
        output_dir: Directory to save import files
        forms: List of form types to generate (None = all available)
        client_id: Optional CCH client ID

    Returns:
        Dict of generated file paths by form type
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    generated = {}
    form_data = parsed_data.get('forms', {})

    # Determine which forms to generate
    if forms is None:
        forms = ['W2', 'INT', 'DIV', 'SSA', 'MTG', 'R', 'B', 'NEC', 'G', 'MISC', 'K', 'SLI', 'SA', 'K1', 'T', 'Q']

    # W-2
    if 'W2' in forms:
        w2_data = form_data.get('W-2', [])
        w2_data = filter_quality_entries(w2_data, 'W-2')
        if w2_data:
            filename = get_import_filename("W2", w2_data)
            path = output_dir / filename
            generate_w2_import(w2_data, path, client_id)
            generated['W-2'] = path
        else:
            print("  No W-2 data to import")

    # 1099-INT
    if 'INT' in forms:
        int_data = form_data.get('IRS-1099INT', []) or form_data.get('1099-INT', [])
        int_data = filter_quality_entries(int_data, '1099-INT')
        if int_data:
            filename = get_import_filename("1099INT", int_data)
            path = output_dir / filename
            generate_1099int_import(int_data, path, client_id)
            generated['1099-INT'] = path
        else:
            print("  No 1099-INT data to import")

    # 1099-DIV
    if 'DIV' in forms:
        div_data = form_data.get('IRS-1099DIV', []) or form_data.get('1099-DIV', [])
        div_data = filter_quality_entries(div_data, '1099-DIV')
        if div_data:
            filename = get_import_filename("1099DIV", div_data)
            path = output_dir / filename
            generate_1099div_import(div_data, path, client_id)
            generated['1099-DIV'] = path
        else:
            print("  No 1099-DIV data to import")

    # SSA-1099
    if 'SSA' in forms:
        ssa_data = form_data.get('SSA-1099', [])
        ssa_data = filter_quality_entries(ssa_data, 'SSA-1099')
        if ssa_data:
            filename = get_import_filename("SSA1099", ssa_data)
            path = output_dir / filename
            generate_ssa1099_import(ssa_data, path, client_id)
            generated['SSA-1099'] = path
        else:
            print("  No SSA-1099 data to import")

    # 1098 Mortgage Interest
    if 'MTG' in forms:
        mtg_data = form_data.get('1098', [])
        mtg_data = filter_quality_entries(mtg_data, '1098')
        if mtg_data:
            filename = get_import_filename("1098", mtg_data)
            path = output_dir / filename
            generate_1098_import(mtg_data, path, client_id)
            generated['1098'] = path
        else:
            print("  No 1098 data to import")

    # 1099-R Retirement Distributions
    if 'R' in forms:
        r_data = form_data.get('IRS-1099R', []) or form_data.get('1099-R', [])
        r_data = filter_quality_entries(r_data, '1099-R')
        if r_data:
            filename = get_import_filename("1099R", r_data)
            path = output_dir / filename
            generate_1099r_import(r_data, path, client_id)
            generated['1099-R'] = path
        else:
            print("  No 1099-R data to import")

    # 1099-B Capital Gains
    if 'B' in forms:
        b_data = form_data.get('1099-B', [])
        b_data = filter_quality_entries(b_data, '1099-B')
        if b_data:
            filename = get_import_filename("1099B", b_data)
            path = output_dir / filename
            generate_1099b_import(b_data, path, client_id)
            generated['1099-B'] = path
        else:
            print("  No 1099-B data to import")

    # 1099-NEC Nonemployee Compensation
    if 'NEC' in forms:
        nec_data = form_data.get('1099-NEC', [])
        nec_data = filter_quality_entries(nec_data, '1099-NEC')
        if nec_data:
            filename = get_import_filename("1099NEC", nec_data)
            path = output_dir / filename
            generate_1099nec_import(nec_data, path, client_id)
            generated['1099-NEC'] = path
        else:
            print("  No 1099-NEC data to import")

    # 1099-G Government Payments
    if 'G' in forms:
        g_data = form_data.get('1099-G', [])
        g_data = filter_quality_entries(g_data, '1099-G')
        if g_data:
            filename = get_import_filename("1099G", g_data)
            path = output_dir / filename
            generate_1099g_import(g_data, path, client_id)
            generated['1099-G'] = path
        else:
            print("  No 1099-G data to import")

    # 1099-MISC Miscellaneous
    if 'MISC' in forms:
        misc_data = form_data.get('1099-MISC', [])
        misc_data = filter_quality_entries(misc_data, '1099-MISC')
        if misc_data:
            filename = get_import_filename("1099MISC", misc_data)
            path = output_dir / filename
            generate_1099misc_import(misc_data, path, client_id)
            generated['1099-MISC'] = path
        else:
            print("  No 1099-MISC data to import")

    # 1099-K Payment Card
    if 'K' in forms:
        k_data = form_data.get('1099-K', [])
        k_data = filter_quality_entries(k_data, '1099-K')
        if k_data:
            filename = get_import_filename("1099K", k_data)
            path = output_dir / filename
            generate_1099k_import(k_data, path, client_id)
            generated['1099-K'] = path
        else:
            print("  No 1099-K data to import")

    # 1098-E Student Loan Interest
    if 'SLI' in forms:
        sli_data = form_data.get('1098-E', [])
        sli_data = filter_quality_entries(sli_data, '1098-E')
        if sli_data:
            filename = get_import_filename("1098E", sli_data)
            path = output_dir / filename
            generate_1098e_import(sli_data, path, client_id)
            generated['1098-E'] = path
        else:
            print("  No 1098-E data to import")

    # 1099-SA HSA/MSA
    if 'SA' in forms:
        sa_data = form_data.get('1099-SA', [])
        sa_data = filter_quality_entries(sa_data, '1099-SA')
        if sa_data:
            filename = get_import_filename("1099SA", sa_data)
            path = output_dir / filename
            generate_1099sa_import(sa_data, path, client_id)
            generated['1099-SA'] = path
        else:
            print("  No 1099-SA data to import")

    # K-1 Partnership
    if 'K1' in forms:
        k1_data = form_data.get('K-1', []) or form_data.get('K1-Partnership', [])
        k1_data = filter_quality_entries(k1_data, 'K-1')
        if k1_data:
            filename = get_import_filename("K1", k1_data)
            path = output_dir / filename
            generate_k1_import(k1_data, path, client_id)
            generated['K-1'] = path
        else:
            print("  No K-1 data to import")

    # 1098-T Tuition Statement
    if 'T' in forms:
        t_data = form_data.get('1098-T', [])
        t_data = filter_quality_entries(t_data, '1098-T')
        if t_data:
            filename = get_import_filename("1098T", t_data)
            path = output_dir / filename
            generate_1098t_import(t_data, path, client_id)
            generated['1098-T'] = path
        else:
            print("  No 1098-T data to import")

    # 1099-Q Qualified Education
    if 'Q' in forms:
        q_data = form_data.get('1099-Q', [])
        q_data = filter_quality_entries(q_data, '1099-Q')
        if q_data:
            filename = get_import_filename("1099Q", q_data)
            path = output_dir / filename
            generate_1099q_import(q_data, path, client_id)
            generated['1099-Q'] = path
        else:
            print("  No 1099-Q data to import")

    return generated


def main():
    parser = argparse.ArgumentParser(
        description='Generate CCH ProSystem fx import files from parsed tax documents',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python generate_cch_import.py parsed_data.json
  python generate_cch_import.py parsed_data.json --output-dir "L:\\Client\\2025"
  python generate_cch_import.py parsed_data.json --forms INT,DIV
  python generate_cch_import.py parsed_data.json --client-id "24I:clientname:V1"

Supported form codes:
  INT  - 1099-INT (Interest)
  DIV  - 1099-DIV (Dividends)
  SSA  - SSA-1099 (Social Security)

CCH Import Instructions:
  1. Open CCH ProSystem fx Tax
  2. Open the client return
  3. Navigate to the worksheet (e.g., Interest)
  4. Click Import button
  5. Select the generated Excel file
  6. In Import Wizard, set starting row to 7
  7. Verify column mapping
  8. Import
        """
    )

    parser.add_argument('json_file', help='Parsed data JSON file from parse_source_docs.py')
    parser.add_argument('--output-dir', '-o',
                        help='Output directory (default: same as JSON file)')
    parser.add_argument('--forms', '-f',
                        help='Comma-separated form codes: INT,DIV,SSA (default: all)')
    parser.add_argument('--client-id', '-c',
                        help='CCH client ID (e.g., "24I:clientname:V1")')

    args = parser.parse_args()

    json_path = Path(args.json_file)
    if not json_path.exists():
        print(f"ERROR: File not found: {json_path}")
        sys.exit(1)

    # Load parsed data
    print(f"Loading: {json_path}")
    with open(json_path) as f:
        parsed_data = json.load(f)

    # Determine output directory
    if args.output_dir:
        output_dir = Path(args.output_dir)
    else:
        output_dir = json_path.parent

    # Parse form codes
    forms = None
    if args.forms:
        forms = [f.strip().upper() for f in args.forms.split(',')]

    print(f"\n{'='*60}")
    print("CCH IMPORT FILE GENERATOR")
    print('='*60)
    print(f"Output directory: {output_dir}")
    print()

    # Generate import files
    generated = generate_cch_imports(
        parsed_data,
        output_dir,
        forms=forms,
        client_id=args.client_id
    )

    print(f"\n{'='*60}")
    print("GENERATION COMPLETE")
    print('='*60)
    print(f"Files created: {len(generated)}")
    for form_type, path in generated.items():
        print(f"  {form_type}: {path.name}")

    if generated:
        print("\nNext steps:")
        print("  1. Open CCH ProSystem fx Tax")
        print("  2. Open client return")
        print("  3. Go to appropriate worksheet")
        print("  4. Click Import -> Select file -> Set start row to 7")
        print("  5. Verify mapping -> Import")


if __name__ == '__main__':
    main()
